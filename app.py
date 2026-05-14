import html
import tempfile
import re
from collections import defaultdict
from decimal import Decimal
from pathlib import Path
from time import perf_counter

import pandas as pd
import plotly.express as px
import pdfplumber
import streamlit as st

try:
    import fitz
except ModuleNotFoundError:
    fitz = None

import auditoria_engine as auditoria_io
from auditoria_engine import auditar, format_money_br, linhas_para_dataframe

BRAND_NAME = "FRETE VISION"
BRAND_TAGLINE = "Visão que move resultados"
BRAND_PLATFORM = "Plataforma de Inteligência Logística"

STATUS_DISPLAY_MAP = {"OK por arredondamento": "OK Arred."}
STATUS_ORDER_MAP = {
    "Divergente": 0,
    "Faltante no A": 1,
    "Faltante no B": 2,
    "OK por arredondamento": 3,
    "OK": 4,
}
CONFERENCE_FILTER_LABELS = [
    "Todos",
    "Críticos",
    "Divergentes reais",
    "Diferenças dentro da tolerância",
    "Faltantes",
    "OK sem diferença",
]
VISUAL_DIFF_FILTERS = {
    "R$ 0,00": 0.00,
    "R$ 0,01": 0.01,
    "R$ 0,10": 0.10,
    "R$ 0,50": 0.50,
    "R$ 1,00": 1.00,
    "Personalizado": None,
}
RE_GW_MARGIN_VISUAL = re.compile(r"^\s*0*(\d{4,})\b.*?(-?\d{1,3}(?:\.\d{3})*,\d{2}%)\s*$")
PROCESSING_STEPS = [
    ("Lendo Relatório A", "Validando o arquivo temporário da base ATUA."),
    ("Lendo Relatório B", "Validando o arquivo temporário da base GW."),
    ("Cruzando CTEs", "Comparando os registros entre os dois relatórios."),
    ("Aplicando tolerância", "Classificando diferenças conforme a tolerância configurada."),
    ("Gerando relatório", "Montando a tabela final e preparando as saídas."),
]
MAX_UPLOAD_SIZE_BYTES = 20 * 1024 * 1024
PDF_HEADER_SCAN_BYTES = 1024


st.set_page_config(page_title=BRAND_NAME, layout="wide", initial_sidebar_state="collapsed")

def ui(text):
    if not isinstance(text, str):
        return text
    for source_encoding in ("cp1252", "latin1"):
        try:
            return text.encode(source_encoding).decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            continue
    return text





CSS = """
<style>
:root {
    --bg: #f4f0e7;
    --surface: rgba(255, 255, 255, 0.96);
    --surface-soft: #f8f3ea;
    --sidebar: #ffffff;
    --sidebar-top: #f6efe1;
    --text: #13253a;
    --muted: #697587;
    --muted-2: #9ca8b5;
    --line: #e8dfd1;
    --purple: #b6843c;
    --purple-dark: #916226;
    --purple-soft: #f5ede1;
    --purple-ink: #14273f;
    --pink: #c08b41;
    --orange: #cf8f37;
    --blue: #355d88;
    --green: #22c55e;
    --green-soft: #dcfce7;
    --yellow: #f59e0b;
    --yellow-soft: #fef3c7;
    --red: #ef4444;
    --red-soft: #fee2e2;
    --shadow: 0 18px 40px rgba(18, 34, 53, 0.08);
    --shadow-sm: 0 10px 24px rgba(18, 34, 53, 0.06);
    --process-title: #111827;
    --process-subtitle: #475569;
    --process-active: #111827;
    --process-done: #16a34a;
    --process-pending: #94a3b8;
    --process-border: #e5e7eb;
    --process-surface: #ffffff;
    --process-surface-soft: #f8fafc;
    --process-progress: #2563eb;
    --process-progress-track: #dbeafe;
}

html, body, [class*="css"] {
    font-family: Manrope, "Segoe UI", sans-serif !important;
    color: var(--text) !important;
}

.stApp {
    background:
        radial-gradient(circle at top right, rgba(199, 149, 78, 0.14), transparent 28%),
        radial-gradient(circle at bottom left, rgba(20, 39, 63, 0.08), transparent 26%),
        linear-gradient(180deg, #f9f5ee 0%, #f4f0e7 100%) !important;
}

header, footer, #MainMenu {
    display: none !important;
}

.block-container {
    max-width: 100% !important;
    padding: 0 28px 42px 28px !important;
}

h1, h2, h3, h4, p {
    font-family: Manrope, "Segoe UI", sans-serif !important;
    letter-spacing: 0 !important;
    text-transform: none !important;
}

[data-testid="stSidebar"] {
    background: var(--sidebar) !important;
    border-right: 1px solid #ede4d7 !important;
    box-shadow: 8px 0 22px rgba(20, 39, 63, 0.04);
}

[data-testid="stSidebar"] > div:first-child {
    padding: 0 !important;
}

.sidebar-brand {
    min-height: 92px;
    padding: 16px 20px;
    background: linear-gradient(135deg, #10233a 0%, #18304b 100%);
    display: flex;
    align-items: center;
    gap: 12px;
    border-bottom: 1px solid #d7c8b2;
}

.logo-mark {
    width: 54px;
    height: 54px;
    display: flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
}

.logo-mark svg {
    width: 100%;
    height: 100%;
    display: block;
    filter: drop-shadow(0 10px 18px rgba(0, 0, 0, 0.16));
}

.brand-name {
    font-size: 1.08rem;
    line-height: 1;
    font-weight: 800;
    color: #ffffff;
    letter-spacing: 0.18em !important;
}

.brand-caption {
    margin-top: 4px;
    font-size: 0.72rem;
    font-weight: 700;
    color: #e6c78f;
    letter-spacing: 0.22em !important;
    text-transform: uppercase !important;
}

.sidebar-section {
    padding: 22px 10px 0 10px;
}

[data-testid="stSidebar"] .stRadio > label {
    display: none !important;
}

[data-testid="stSidebar"] [role="radiogroup"] {
    display: flex !important;
    flex-direction: column !important;
    gap: 8px !important;
}

[data-testid="stSidebar"] [role="radiogroup"] label {
    min-height: 46px !important;
    border-radius: 0 24px 24px 0 !important;
    border: 0 !important;
    border-left: 4px solid transparent !important;
    padding: 0 14px !important;
    color: #4f5f72 !important;
    font-weight: 700 !important;
    background: transparent !important;
    transition: all 160ms ease !important;
}

[data-testid="stSidebar"] [role="radiogroup"] label:hover {
    color: var(--purple-ink) !important;
    background: #f7f1e7 !important;
}

[data-testid="stSidebar"] [role="radiogroup"] label[data-checked="true"] {
    color: var(--purple-ink) !important;
    background: var(--purple-soft) !important;
    border-left-color: var(--purple-dark) !important;
    box-shadow: inset 0 0 0 1px #e3d1b6;
}

[data-testid="stSidebar"] input[type="radio"] {
    display: none !important;
}

[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
    margin: 0 !important;
}

.sidebar-profile {
    margin: 28px 18px 0 18px;
    padding: 16px;
    border-top: 1px solid var(--line);
}

.user-chip {
    display: flex;
    align-items: center;
    gap: 10px;
}

.avatar {
    width: 38px;
    height: 38px;
    border-radius: 50%;
    display: grid;
    place-items: center;
    background: linear-gradient(135deg, #10233a, #274565);
    border: 3px solid #efe4d2;
    color: #e7c58c;
    font-weight: 800;
}

.user-name {
    color: var(--text);
    font-size: 0.88rem;
    font-weight: 800;
}

.user-email {
    color: var(--muted);
    font-size: 0.72rem;
    margin-top: 2px;
}

.progress-mini {
    margin-top: 16px;
}

.progress-label {
    display: flex;
    justify-content: space-between;
    color: var(--muted);
    font-size: 0.72rem;
    font-weight: 700;
}

.progress-bar {
    margin-top: 8px;
    height: 7px;
    border-radius: 99px;
    background: #ede5d7;
    overflow: hidden;
}

.progress-fill {
    width: 78%;
    height: 100%;
    background: linear-gradient(90deg, #b6843c, #e8c88d);
}

.topbar {
    height: 70px;
    margin: 0 -28px 24px -28px;
    padding: 0 28px;
    background: rgba(255, 250, 244, 0.84);
    border-bottom: 1px solid #ebdfcf;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 24px;
    backdrop-filter: blur(12px);
}

.topbar-left {
    display: flex;
    align-items: center;
    gap: 16px;
}

.hamburger {
    width: 36px;
    height: 36px;
    border-radius: 12px;
    display: grid;
    place-items: center;
    color: #657387;
}

.page-name {
    font-size: 1.05rem;
    font-weight: 800;
    color: #16273e;
}

.topbar-right {
    display: flex;
    align-items: center;
    gap: 14px;
}

.search-box {
    width: min(330px, 30vw);
    height: 44px;
    border-radius: 24px;
    background: linear-gradient(135deg, #fffdfa 0%, #f7efe2 100%);
    border: 1px solid #e6d7c2;
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 0 16px;
    color: #516173;
    font-size: 0.86rem;
    font-weight: 800;
    box-shadow: 0 6px 14px rgba(56, 49, 85, 0.04);
}

.top-icon {
    width: 36px;
    height: 36px;
    border-radius: 12px;
    display: grid;
    place-items: center;
    color: #46586e;
    background: rgba(255, 255, 255, 0.92);
    border: 1px solid #ece0d0;
    position: relative;
    box-shadow: 0 6px 12px rgba(20, 39, 63, 0.04);
}

.top-icon .dot {
    position: absolute;
    top: 2px;
    right: 0;
    min-width: 15px;
    height: 15px;
    border-radius: 99px;
    background: #b6843c;
    color: white;
    font-size: 0.55rem;
    display: grid;
    place-items: center;
    font-weight: 800;
}

.top-avatar {
    width: 42px;
    height: 42px;
    border-radius: 50%;
    background: linear-gradient(135deg, #10233a, #274565);
    border: 3px solid #efe4d2;
    display: grid;
    place-items: center;
    color: #e7c58c;
    font-weight: 900;
}

.breadcrumb {
    min-height: 60px;
    padding: 0 26px;
    border-radius: 10px;
    background: linear-gradient(135deg, rgba(255, 255, 255, 0.98), rgba(247, 241, 232, 0.96));
    display: flex;
    align-items: center;
    color: #778294;
    font-size: 0.98rem;
    box-shadow: var(--shadow-sm);
    margin-bottom: 24px;
}

.breadcrumb b {
    color: #1a2b41;
}

.breadcrumb span {
    color: var(--pink);
    margin-left: 8px;
}

.dashboard-grid {
    display: grid;
    grid-template-columns: repeat(12, minmax(0, 1fr));
    gap: 24px;
}

.panel {
    grid-column: span 12;
    background: var(--surface);
    border-radius: 14px;
    box-shadow: var(--shadow);
    overflow: hidden;
    border: 1px solid rgba(232, 223, 209, 0.84);
}

.panel.span-3 { grid-column: span 3; }
.panel.span-4 { grid-column: span 4; }
.panel.span-6 { grid-column: span 6; }
.panel.span-8 { grid-column: span 8; }

.panel-title {
    min-height: 70px;
    padding: 0 22px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 1px solid #eeedf3;
}

.panel-title h3 {
    margin: 0;
    color: #16273e;
    font-size: 1rem;
    font-weight: 800;
}

.panel-subtitle {
    color: var(--muted);
    font-size: 0.78rem;
    font-weight: 600;
    margin-top: 3px;
}

.panel-body {
    padding: 22px;
}

.hero-shell {
    overflow: visible;
}

.hero-layout {
    display: grid;
    grid-template-columns: minmax(0, 1.15fr) minmax(300px, 0.85fr);
    gap: 26px;
    align-items: stretch;
}

.hero-copy {
    padding-right: 8px;
}

.hero-kicker {
    display: inline-flex;
    align-items: center;
    gap: 10px;
    min-height: 34px;
    padding: 0 14px;
    border-radius: 999px;
    border: 1px solid #e4d2b5;
    background: linear-gradient(135deg, #fffdf9 0%, #f7efe3 100%);
    color: #8c6631;
    font-size: 0.76rem;
    font-weight: 800;
    letter-spacing: 0.16em;
    text-transform: uppercase;
}

.hero-copy h1 {
    margin: 18px 0 12px 0;
    color: #10233a;
    font-family: "Cormorant Garamond", Georgia, serif !important;
    font-size: clamp(2.8rem, 4vw, 4.6rem);
    line-height: 0.95;
    font-weight: 700;
}

.hero-copy p {
    max-width: 720px;
    color: #5f6d80;
    line-height: 1.8;
    margin: 0;
    font-size: 1rem;
}

.hero-actions {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 22px;
}

.hero-action {
    min-height: 42px;
    padding: 0 16px;
    border-radius: 999px;
    border: 1px solid #ded2c2;
    background: rgba(255, 255, 255, 0.82);
    color: #17304a;
    font-size: 0.82rem;
    font-weight: 800;
    display: inline-flex;
    align-items: center;
}

.hero-action.primary {
    background: linear-gradient(135deg, #10233a 0%, #18304b 100%);
    border-color: #10233a;
    color: #f6e2ba;
    box-shadow: 0 16px 26px rgba(16, 35, 58, 0.18);
}

.hero-stats {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 14px;
    margin-top: 28px;
}

.hero-stat {
    padding: 16px;
    border-radius: 18px;
    background: linear-gradient(180deg, rgba(255, 255, 255, 0.94) 0%, rgba(248, 241, 231, 0.9) 100%);
    border: 1px solid #e8ddcd;
}

.hero-stat b {
    display: block;
    color: #10233a;
    font-size: 1.45rem;
    line-height: 1;
    margin-bottom: 8px;
}

.hero-stat span {
    color: #677486;
    font-size: 0.8rem;
    line-height: 1.5;
    font-weight: 700;
}

.hero-visual {
    min-height: 100%;
    border-radius: 20px;
    background: linear-gradient(160deg, #fffdf9 0%, #f6efe3 54%, #eff2f5 100%);
    display: grid;
    place-items: stretch;
    border: 1px solid #eadfce;
    box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.78);
}

.hero-brand-card {
    padding: 24px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    gap: 22px;
}

.fv-signature {
    display: grid;
    justify-items: center;
    gap: 12px;
}

.fv-symbol {
    width: min(100%, 210px);
}

.fv-symbol svg {
    width: 100%;
    height: auto;
    display: block;
}

.fv-wordmark {
    text-align: center;
}

.fv-wordmark-main {
    font-size: clamp(1.5rem, 2vw, 2.05rem);
    letter-spacing: 0.28em;
    color: #0f233a;
    font-weight: 700;
}

.fv-wordmark-main strong {
    color: #b6843c;
    font-weight: 700;
}

.fv-wordmark-tagline {
    margin-top: 10px;
    display: inline-flex;
    align-items: center;
    gap: 12px;
    color: #6c788a;
    font-size: 0.7rem;
    font-weight: 800;
    letter-spacing: 0.28em;
    text-transform: uppercase;
}

.fv-wordmark-tagline span {
    width: 42px;
    height: 1px;
    background: linear-gradient(90deg, transparent, #cda369, transparent);
}

.hero-proof-grid,
.feature-strip {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 12px;
}

.hero-proof,
.feature-card {
    padding: 16px;
    border-radius: 16px;
    border: 1px solid #eadfce;
    background: rgba(255, 255, 255, 0.74);
}

.hero-proof small,
.feature-card small {
    display: block;
    color: #9a7440;
    font-size: 0.68rem;
    letter-spacing: 0.16em;
    text-transform: uppercase;
    font-weight: 800;
    margin-bottom: 8px;
}

.hero-proof b,
.feature-card b {
    display: block;
    color: #13253a;
    font-size: 0.95rem;
    margin-bottom: 6px;
}

.feature-strip {
    margin-top: 20px;
    grid-template-columns: repeat(4, minmax(0, 1fr));
}

.feature-card span,
.hero-proof span {
    color: #677486;
    font-size: 0.78rem;
    line-height: 1.55;
}

.upload-shell {
    border: 1px dashed #dccdaf;
    border-radius: 14px;
    background: #fffcf7;
    padding: 18px;
}

.upload-head {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 14px;
}

.small-icon {
    width: 38px;
    height: 38px;
    border-radius: 12px;
    display: grid;
    place-items: center;
    color: var(--purple-dark);
    background: #f5ede1;
}

.upload-name {
    color: var(--text);
    font-weight: 800;
    font-size: 0.95rem;
}

.upload-help {
    color: var(--muted);
    font-size: 0.78rem;
    margin-top: 2px;
}

[data-testid="stFileUploader"] {
    background: #ffffff !important;
    border: 1px solid #eadfce !important;
    border-radius: 14px !important;
    padding: 14px !important;
}

[data-testid="stFileUploader"] button {
    border-radius: 10px !important;
    border: 0 !important;
    background: #10233a !important;
    color: white !important;
    font-weight: 800 !important;
}

.loaded-file {
    margin-top: 12px;
    padding: 11px 12px;
    border-radius: 9px;
    background: #edfff4;
    color: #15803d;
    border: 1px solid #c8f6d8;
    font-size: 0.78rem;
    font-weight: 800;
}

.loaded-file span {
    display: block;
    margin-top: 3px;
    color: #166534;
    font-weight: 600;
}

[data-testid="stRadio"] [role="radiogroup"] {
    display: flex !important;
    flex-wrap: wrap !important;
    gap: 8px !important;
    background: #f4ecdf !important;
    border: 1px solid #e3d1b6 !important;
    border-radius: 14px !important;
    padding: 6px !important;
}

[data-testid="stRadio"] [role="radiogroup"] label {
    border-radius: 10px !important;
    border: 1px solid #decdb3 !important;
    color: #19304a !important;
    padding: 8px 14px !important;
    font-weight: 900 !important;
    background: rgba(255, 255, 255, 0.92) !important;
}

[data-testid="stRadio"] [role="radiogroup"] label * {
    color: inherit !important;
    font-weight: inherit !important;
}

[data-testid="stRadio"] [role="radiogroup"] label[data-checked="true"] {
    background: #10233a !important;
    color: #f7e0b6 !important;
    box-shadow: 0 10px 18px rgba(16, 35, 58, 0.22);
    border-color: #10233a !important;
}

.stButton > button[kind="primary"] {
    height: 50px !important;
    border-radius: 12px !important;
    border: 0 !important;
    background: linear-gradient(135deg, #10233a 0%, #18304b 100%) !important;
    color: white !important;
    font-weight: 900 !important;
    box-shadow: 0 12px 22px rgba(16, 35, 58, 0.25) !important;
    transition: all 160ms ease !important;
}

.stButton > button[kind="primary"]:hover {
    background: #18304b !important;
    transform: translateY(-1px);
}

.stButton > button, .stDownloadButton > button {
    border-radius: 9px !important;
    font-weight: 800 !important;
}

.kpi-value {
    color: #252638;
    font-size: clamp(1.6rem, 2.4vw, 2.3rem);
    line-height: 1;
    font-weight: 900;
    white-space: nowrap;
}

.kpi-value.money {
    font-size: clamp(1rem, 1.25vw, 1.35rem);
}

.kpi-label {
    color: var(--muted);
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    font-weight: 900;
    margin-bottom: 14px;
}

.tone-green { color: var(--green); }
.tone-blue { color: var(--blue); }
.tone-purple { color: var(--purple); }
.tone-orange { color: var(--orange); }
.tone-red { color: var(--red); }

.explain {
    color: var(--muted);
    font-size: 0.86rem;
    line-height: 1.7;
}

.explain b {
    color: var(--text);
}

.impact-line {
    text-align: right;
    color: var(--muted);
    font-size: 0.9rem;
}

.impact-line b {
    color: var(--text);
}

.status-ok { background: var(--green-soft); color: #166534; }
.status-round { background: #e8f3ff; color: #075985; }
.status-div { background: var(--red-soft); color: #991b1b; }
.status-miss { background: var(--yellow-soft); color: #92400e; }

[data-testid="stVerticalBlockBorderWrapper"] {
    border: 1px solid rgba(233, 231, 241, 0.92) !important;
    border-radius: 14px !important;
    background: #ffffff !important;
    box-shadow: var(--shadow) !important;
    padding: 6px !important;
}

.stTextInput label, .stNumberInput label, .stSelectbox label {
    color: #64748b !important;
    font-weight: 700 !important;
}

.stTextInput input,
.stNumberInput input {
    min-height: 42px !important;
    border-radius: 10px !important;
    border: 1px solid #d7dcee !important;
    background: #ffffff !important;
    color: #0f172a !important;
    caret-color: #0f172a !important;
    box-shadow: none !important;
}

.stTextInput input::placeholder,
.stNumberInput input::placeholder {
    color: #94a3b8 !important;
    opacity: 1 !important;
}

.stTextInput input:focus,
.stNumberInput input:focus {
    border-color: var(--purple) !important;
    box-shadow: 0 0 0 3px rgba(120, 98, 214, 0.14) !important;
}

.stSelectbox [data-baseweb="select"] {
    min-height: 42px !important;
    border-radius: 10px !important;
    border: 1px solid #d7dcee !important;
    background: #ffffff !important;
    color: #0f172a !important;
    box-shadow: none !important;
}

.stSelectbox [data-baseweb="select"] > div,
.stSelectbox [data-baseweb="select"] > div > div,
.stSelectbox [data-baseweb="select"] > div > div > div,
.stSelectbox [data-baseweb="select"] span,
.stSelectbox [data-baseweb="select"] input,
.stSelectbox [data-baseweb="select"] * {
    background: #ffffff !important;
    color: #0f172a !important;
    fill: #64748b !important;
}

div[data-baseweb="popover"] {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
    box-shadow: 0 18px 38px rgba(15, 23, 42, 0.12) !important;
}

div[data-baseweb="popover"],
div[data-baseweb="popover"] > div,
div[data-baseweb="popover"] ul,
div[data-baseweb="popover"] li,
div[data-baseweb="popover"] * {
    background: #ffffff !important;
    color: #0f172a !important;
}

[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    border: 1px dashed #d9cef8 !important;
    border-radius: 14px !important;
}

[data-testid="stFileUploaderDropzone"] * {
    color: #64748b !important;
}

[data-testid="stFileUploaderFile"] {
    background: #f8fafc !important;
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
}

[data-testid="stFileUploaderFile"],
[data-testid="stFileUploaderFile"] > div,
[data-testid="stFileUploaderFile"] section,
[data-testid="stFileUploaderFile"] button,
[data-testid="stFileUploaderFile"] * {
    color: #0f172a !important;
    background: transparent !important;
}

div[data-baseweb="input"] input,
div[data-baseweb="base-input"] input,
div[data-baseweb="select"] input {
    color: #0f172a !important;
    caret-color: #0f172a !important;
}

.stNumberInput [data-baseweb="input"],
.stNumberInput [data-baseweb="base-input"],
.stNumberInput [data-baseweb="input"] > div,
.stNumberInput [data-baseweb="base-input"] > div {
    background: #ffffff !important;
    color: #0f172a !important;
}

.stNumberInput button,
.stNumberInput [data-baseweb="input"] button,
.stNumberInput [data-baseweb="base-input"] button {
    background: #ffffff !important;
    color: #475569 !important;
    border: 1px solid #d7dcee !important;
    box-shadow: none !important;
}

.table-shell {
    margin-top: 12px;
    overflow: auto;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    background: #ffffff;
}

.audit-results-shell {
    max-height: 68vh;
    min-height: 320px;
    overscroll-behavior: contain;
    scrollbar-gutter: stable both-edges;
}

.audit-table {
    width: 100%;
    min-width: 1420px;
    border-collapse: separate;
    border-spacing: 0;
}

.audit-table thead th {
    background: #f8fafc;
    color: #64748b;
    font-size: 0.74rem;
    font-weight: 900;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    padding: 14px 16px;
    text-align: left;
    border-bottom: 1px solid #e2e8f0;
    white-space: nowrap;
    position: sticky;
    top: 0;
    z-index: 2;
}

.audit-table tbody td {
    padding: 14px 16px;
    border-bottom: 1px solid #eef2f7;
    color: #0f172a;
    background: #ffffff;
    font-size: 0.9rem;
    vertical-align: middle;
}

.audit-table tbody tr:nth-child(even) td {
    background: #fbfcfe;
}

.audit-table tbody tr:hover td {
    background: #f8fbff;
}

.audit-table tbody tr.row-missing td {
    background: #fff7ed !important;
    color: #9a3412;
}

.audit-table tbody tr.row-missing:hover td {
    background: #ffedd5 !important;
}

.audit-table .cell-cte {
    font-weight: 800;
}

.audit-table .cell-money {
    text-align: right;
    white-space: nowrap;
    font-variant-numeric: tabular-nums;
}

.audit-table .cell-observation {
    min-width: 280px;
}

.audit-table .cell-diff-critical {
    background: #fee2e2 !important;
    color: #991b1b !important;
    font-weight: 800;
}

.audit-table .cell-diff-tolerance {
    background: #e8f3ff !important;
    color: #1d4ed8 !important;
    font-weight: 800;
}

.audit-table .cell-empty {
    text-align: center;
    color: #64748b;
    padding: 22px;
}

.table-scroll-hint {
    margin-top: 10px;
    color: #64748b;
    font-size: 0.82rem;
}

.detail-grid {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 12px;
    margin: 6px 0 18px 0;
}

.detail-metric {
    padding: 14px;
    border-radius: 12px;
    background: linear-gradient(180deg, #fffdf7 0%, #f6efe3 100%);
    border: 1px solid #eadfce;
}

.detail-metric small {
    display: block;
    color: #7a6646;
    font-size: 0.7rem;
    font-weight: 900;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}

.detail-metric b {
    display: block;
    margin-top: 6px;
    color: #10233a;
    font-size: 1.3rem;
}

.conference-note {
    margin-top: -6px;
    padding: 12px 14px;
    border-radius: 12px;
    background: #f8fbff;
    border: 1px solid #dbe6f4;
    color: #516173;
    font-size: 0.84rem;
    line-height: 1.6;
}

.table-badge {
    display: inline-flex;
    align-items: center;
    border-radius: 999px;
    padding: 6px 10px;
    font-size: 0.78rem;
    font-weight: 800;
    white-space: nowrap;
}

.table-badge.ok {
    background: #dcfce7;
    color: #166534;
}

.table-badge.round {
    background: #e8f3ff;
    color: #075985;
}

.table-badge.div {
    background: #fee2e2;
    color: #991b1b;
}

.table-badge.miss {
    background: #fef3c7;
    color: #92400e;
}

.export-card h4 {
    margin: 12px 0 4px 0;
    color: var(--text);
    font-size: 1rem;
    font-weight: 900;
}

.export-card p {
    color: var(--muted);
    margin: 0 0 14px 0;
    font-size: 0.82rem;
}

.history-grid {
    display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 12px;
    margin-top: 14px;
}

.history-metric {
    background: #faf8ff;
    border: 1px solid #eeeaf6;
    border-radius: 10px;
    padding: 12px;
}

.history-metric small {
    display: block;
    color: var(--muted);
    font-size: 0.7rem;
    font-weight: 900;
    text-transform: uppercase;
}

.history-metric b {
    display: block;
    color: var(--text);
    margin-top: 4px;
}

.ops-note {
    margin-top: 14px;
    padding: 12px 14px;
    border-radius: 12px;
    background: #f8fafc;
    border: 1px solid #e7edf6;
    color: #64748b;
    font-size: 0.8rem;
    line-height: 1.55;
}

.ops-note b {
    color: #1f2937;
}

[data-baseweb="select"] > div,
[data-baseweb="select"] > div > div,
[data-baseweb="select"] > div > div > div,
[data-baseweb="input"] > div,
[data-baseweb="base-input"] > div {
    background: #ffffff !important;
    color: #0f172a !important;
}

[role="listbox"],
[role="listbox"] > li,
[role="option"] {
    background: #ffffff !important;
    color: #0f172a !important;
}

[role="listbox"] > li:hover,
[role="option"]:hover {
    background: #f8fafc !important;
}


.upload-card {
    display: block;
}

.upload-shell {
    border: 0;
    border-radius: 0;
    background: transparent;
    padding: 0;
    margin-bottom: 12px;
}

.upload-head {
    align-items: flex-start;
    gap: 14px;
    margin-bottom: 0;
}

.upload-copy {
    flex: 1;
    min-width: 0;
}

.upload-title-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
}

.small-icon {
    width: 40px;
    height: 40px;
    border-radius: 14px;
    display: grid;
    place-items: center;
    color: var(--purple-dark);
    background: linear-gradient(180deg, #f6f1ff 0%, #efe7ff 100%);
    border: 1px solid #e5dcfb;
    flex-shrink: 0;
}

.upload-name {
    color: #202236;
    font-weight: 800;
    font-size: 1rem;
    line-height: 1.2;
}

.upload-help {
    color: #6f778b;
    font-size: 0.8rem;
    line-height: 1.45;
    margin-top: 6px;
}

.upload-badge {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-height: 24px;
    padding: 0 10px;
    border-radius: 999px;
    border: 1px solid #d9e2f0;
    background: #ffffff;
    color: #64748b;
    font-size: 0.72rem;
    font-weight: 800;
    white-space: nowrap;
}

.upload-badge.loaded {
    border-color: #cfe8d9;
    background: #f3fbf7;
    color: #15803d;
}

.upload-inline-status {
    margin-top: 10px;
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 10px 12px;
    border-radius: 12px;
    border: 1px solid #dce8f5;
    background: #f8fbff;
    color: #475569;
    font-size: 0.79rem;
}

.upload-inline-status strong {
    color: #1f2937;
    font-weight: 800;
}

.upload-inline-status span:last-child {
    color: #64748b;
}

.upload-inline-dot {
    width: 8px;
    height: 8px;
    border-radius: 999px;
    background: #16a34a;
    flex-shrink: 0;
}

.panel-title {
    min-height: 64px;
    padding: 0 20px;
}

.panel-title h3 {
    font-size: 0.98rem;
}

.panel-subtitle {
    font-size: 0.77rem;
}

.panel-body {
    padding: 18px 20px;
}

.search-box {
    min-width: 210px;
    height: 40px;
    justify-content: flex-start;
    gap: 10px;
    padding: 0 14px;
    color: #4f5f72;
    font-size: 0.82rem;
    font-weight: 800;
    background: linear-gradient(135deg, #fffdf9 0%, #f6efe2 100%);
    border: 1px solid #e5d5be;
    box-shadow: 0 6px 14px rgba(20, 39, 63, 0.05);
}

.search-box span {
    white-space: nowrap;
}

.search-box svg {
    color: #b6843c;
}

[data-testid="stFileUploader"] {
    background: transparent !important;
    border: 0 !important;
    border-radius: 0 !important;
    padding: 0 !important;
}

[data-testid="stFileUploader"] button {
    height: 38px !important;
    border-radius: 10px !important;
    border: 1px solid #dcc9ab !important;
    background: #ffffff !important;
    color: var(--purple-dark) !important;
    box-shadow: none !important;
    font-weight: 800 !important;
}

[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    border: 1px dashed #cbd5e1 !important;
    border-radius: 14px !important;
    min-height: 76px !important;
    padding: 10px 12px !important;
}

[data-testid="stFileUploaderDropzone"] > div {
    padding: 0 !important;
}

[data-testid="stFileUploaderDropzone"] * {
    color: #475569 !important;
}

[data-testid="stFileUploaderFile"] {
    margin-top: 10px !important;
    background: #f8fafc !important;
    border: 1px solid #dbe1ea !important;
    border-radius: 12px !important;
    padding: 10px 12px !important;
    box-shadow: none !important;
}

[data-testid="stFileUploaderFile"] * {
    color: #334155 !important;
}

[data-testid="stFileUploaderFile"] button {
    border: 0 !important;
    background: transparent !important;
    color: #64748b !important;
    box-shadow: none !important;
}

.ops-grid {
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 10px;
    margin-top: 10px;
}

.history-metric {
    background: linear-gradient(180deg, #fffcf7 0%, #f7f0e5 100%);
    border: 1px solid #ece0d0;
    border-radius: 12px;
    padding: 12px;
}

.ops-note {
    margin-top: 12px;
    padding: 12px 14px;
    border-radius: 12px;
    background: linear-gradient(180deg, #fffdf9 0%, #f6f8fb 100%);
    border: 1px solid #e4d9ca;
    color: #64748b;
    font-size: 0.79rem;
    line-height: 1.6;
}

.ops-note b {
    color: #1f2937;
}

.processing-shell {
    background: var(--process-surface);
    border: 1px solid var(--process-border);
    border-radius: 20px;
    box-shadow: 0 16px 30px rgba(15, 23, 42, 0.06);
    overflow: hidden;
}

.processing-head {
    padding: 20px 22px 16px;
    border-bottom: 1px solid #eef2f7;
    background: var(--process-surface);
}

.processing-kicker {
    display: inline-block;
    margin-bottom: 8px;
    color: var(--process-subtitle) !important;
    font-size: 0.72rem;
    font-weight: 800;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}

.processing-title {
    margin: 0;
    color: var(--process-title) !important;
    font-size: 1.08rem;
    font-weight: 900;
}

.processing-subtitle {
    margin-top: 6px;
    color: var(--process-subtitle) !important;
    font-size: 0.83rem;
    line-height: 1.6;
}

.processing-stage-note {
    margin-top: 14px;
    display: inline-flex;
    align-items: center;
    gap: 10px;
    padding: 10px 14px;
    border-radius: 999px;
    border: 1px solid var(--process-border);
    background: var(--process-surface-soft);
    color: var(--process-active) !important;
    font-size: 0.78rem;
    font-weight: 800;
}

.processing-stage-note span:last-child {
    color: var(--process-active) !important;
}

.processing-stage-dot {
    width: 10px;
    height: 10px;
    border-radius: 999px;
    background: var(--process-progress);
    box-shadow: 0 0 0 5px rgba(37, 99, 235, 0.14);
    flex-shrink: 0;
}

.processing-grid {
    display: grid;
    grid-template-columns: repeat(5, minmax(0, 1fr));
    gap: 12px;
    padding: 18px 22px 22px;
    background: var(--process-surface);
}

.processing-step {
    min-height: 112px;
    padding: 14px;
    border-radius: 16px;
    border: 1px solid var(--process-border);
    background: var(--process-surface);
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    gap: 12px;
}

.processing-step-top {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 10px;
}

.processing-step-state {
    color: inherit;
    font-size: 0.69rem;
    font-weight: 800;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}

.processing-step-index {
    width: 30px;
    height: 30px;
    border-radius: 999px;
    border: 1px solid currentColor;
    display: grid;
    place-items: center;
    font-size: 0.8rem;
    font-weight: 900;
    line-height: 1;
}

.processing-step-title {
    color: inherit !important;
    font-size: 0.95rem;
    font-weight: 800;
    line-height: 1.35;
}

.processing-step-copy {
    color: inherit !important;
    font-size: 0.78rem;
    line-height: 1.55;
}

.processing-step.step-done {
    background: #f0fdf4;
    border-color: #bbf7d0;
    color: var(--process-done);
}

.processing-step.step-done .processing-step-index {
    background: #dcfce7;
}

.processing-step.step-active {
    background: var(--process-surface-soft);
    border-color: #cbd5e1;
    color: var(--process-active);
    box-shadow: inset 0 0 0 1px rgba(37, 99, 235, 0.12);
}

.processing-step.step-active * {
    color: var(--process-active) !important;
}

.processing-step.step-active .processing-step-index {
    background: #eff6ff;
    color: var(--process-progress);
}

.processing-step.step-pending {
    background: var(--process-surface);
    border-color: var(--process-border);
    color: var(--process-pending);
}

.processing-step.step-pending .processing-step-index {
    background: var(--process-surface-soft);
}

div[data-testid="stProgressBar"] {
    margin-top: 16px;
}

div[data-testid="stProgressBar"] > div {
    background: var(--process-progress-track) !important;
}

div[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #2563eb 0%, #16a34a 100%) !important;
}

[data-testid="stProgressBar"] + div,
[data-testid="stProgressBar"] + div * {
    color: var(--process-subtitle) !important;
}

[data-testid="stExpander"] {
    border: 1px solid var(--process-border) !important;
    border-radius: 16px !important;
    background: #ffffff !important;
    box-shadow: 0 12px 24px rgba(15, 23, 42, 0.04) !important;
    overflow: hidden !important;
}

[data-testid="stExpander"] summary {
    background: #ffffff !important;
    color: var(--process-title) !important;
    padding-top: 0.95rem !important;
    padding-bottom: 0.95rem !important;
}

[data-testid="stExpander"] summary:hover {
    background: #f8fafc !important;
}

[data-testid="stExpander"] summary * {
    color: var(--process-title) !important;
    font-weight: 800 !important;
}

[data-testid="stExpander"] details[open] summary {
    border-bottom: 1px solid var(--process-border) !important;
}

[data-testid="stExpanderDetails"] {
    background: #ffffff !important;
    color: var(--process-title) !important;
}

[data-testid="stExpanderDetails"] * {
    color: inherit !important;
}

[data-testid="stAlertContainer"] * {
    color: inherit !important;
}

@media (max-width: 1300px) {
    .panel.span-3 { grid-column: span 6; }
    .panel.span-4 { grid-column: span 6; }
    .panel.span-6, .panel.span-8 { grid-column: span 12; }
    .processing-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
}

@media (max-width: 760px) {
    .block-container { padding: 0 14px 32px 14px !important; }
    .topbar { margin-left: -14px; margin-right: -14px; padding: 0 14px; }
    .search-box, .top-icon { display: none; }
    .panel.span-3, .panel.span-4, .panel.span-6, .panel.span-8 { grid-column: span 12; }
    .history-grid { grid-template-columns: 1fr; }
    .detail-grid { grid-template-columns: 1fr; }
    .hero-layout,
    .hero-stats,
    .hero-proof-grid,
    .feature-strip { grid-template-columns: 1fr; }
    .hero-copy h1 { font-size: clamp(2.4rem, 12vw, 3.2rem); }
    .fv-wordmark-main { letter-spacing: 0.14em; }
    .processing-grid { grid-template-columns: 1fr; }
}
</style>
"""


st.markdown(CSS, unsafe_allow_html=True)

COMPACT_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Cormorant+Garamond:wght@600;700&display=swap');

:root {
    --bg: #f8fafc;
    --surface: #ffffff;
    --surface-soft: #f8fafc;
    --line: #e5e7eb;
    --text: #111827;
    --muted: #475569;
    --primary: #4f46e5;
    --primary-dark: #4338ca;
    --ok: #16a34a;
    --ok-soft: #f0fdf4;
    --round: #2563eb;
    --round-soft: #eff6ff;
    --div: #dc2626;
    --div-soft: #fef2f2;
    --miss: #f97316;
    --miss-soft: #fff7ed;
    --shadow-sm: 0 18px 38px rgba(15, 23, 42, 0.08);
    --shadow-card: 0 20px 60px rgba(15, 23, 42, 0.08);
    --shadow-lift: 0 26px 70px rgba(79, 70, 229, 0.18);
    --glow: radial-gradient(circle at top, rgba(99, 102, 241, 0.18), rgba(248, 250, 252, 0) 48%);
}

.stApp {
    background:
        radial-gradient(circle at top left, rgba(79, 70, 229, 0.08), transparent 32%),
        radial-gradient(circle at top right, rgba(245, 158, 11, 0.08), transparent 26%),
        linear-gradient(180deg, #f8fafc 0%, #f3f6fb 100%) !important;
    font-family: "Manrope", "Segoe UI", sans-serif !important;
}

[data-testid="stSidebar"],
[data-testid="collapsedControl"] {
    display: none !important;
}

.block-container {
    max-width: 1240px !important;
    padding: 18px 20px 28px 20px !important;
}

div[data-testid="stVerticalBlockBorderWrapper"] {
    background:
        linear-gradient(180deg, rgba(255, 255, 255, 0.98) 0%, rgba(248, 250, 252, 0.96) 100%) !important;
    border: 1px solid rgba(229, 231, 235, 0.9) !important;
    border-radius: 22px !important;
    box-shadow: var(--shadow-sm) !important;
    backdrop-filter: blur(10px) !important;
}

.compact-header-shell {
    position: sticky;
    top: 0;
    z-index: 20;
    background: rgba(248, 250, 252, 0.9);
    backdrop-filter: blur(10px);
    padding-bottom: 10px;
    margin-bottom: 4px;
}

.compact-header {
    position: relative;
    background:
        linear-gradient(135deg, rgba(255,255,255,0.98) 0%, rgba(244,247,255,0.96) 100%);
    border: 1px solid rgba(255,255,255,0.86);
    border-radius: 22px;
    padding: 15px 20px;
    box-shadow: var(--shadow-card);
    overflow: hidden;
}

.compact-header::before {
    content: "";
    position: absolute;
    inset: -20% auto auto -10%;
    width: 240px;
    height: 240px;
    background: var(--glow);
    pointer-events: none;
}

.compact-brand {
    display: flex;
    align-items: center;
    gap: 12px;
}

.compact-brand-mark {
    width: 48px;
    height: 48px;
    flex-shrink: 0;
    filter: drop-shadow(0 10px 18px rgba(16, 35, 58, 0.16));
}

.compact-brand-title {
    color: var(--text);
    font-size: 1.24rem;
    font-weight: 900;
    letter-spacing: 0.03em;
}

.compact-brand-title strong {
    color: #f59e0b;
    font-weight: 900;
}

.compact-brand-subtitle {
    color: var(--muted);
    font-size: 0.78rem;
    font-weight: 700;
    margin-top: 2px;
}

.header-status-pill {
    height: 40px;
    border-radius: 999px;
    border: 1px solid rgba(191, 219, 254, 0.8);
    background: linear-gradient(180deg, #f8fbff 0%, #eaf2ff 100%);
    color: #1d4ed8;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.83rem;
    font-weight: 800;
    padding: 0 14px;
    box-shadow: 0 12px 24px rgba(37, 99, 235, 0.12);
}

.menu-shell {
    margin: 8px 0 18px 0;
}

.menu-caption,
.compact-field-label {
    color: var(--muted);
    font-size: 0.78rem;
    font-weight: 700;
    margin-bottom: 8px;
}

[data-testid="stRadio"] > label,
[data-testid="stSelectbox"] > label,
[data-testid="stNumberInput"] > label,
[data-testid="stTextInput"] > label {
    color: var(--muted) !important;
    font-size: 0.8rem !important;
    font-weight: 700 !important;
}

[data-testid="stRadio"] [role="radiogroup"] {
    display: flex !important;
    flex-wrap: wrap !important;
    gap: 8px !important;
    padding: 0 !important;
    border: 0 !important;
    background: transparent !important;
}

[data-testid="stRadio"] [role="radiogroup"] label {
    border-radius: 12px !important;
    border: 1px solid rgba(226, 232, 240, 0.95) !important;
    color: var(--text) !important;
    padding: 8px 14px !important;
    min-height: 38px !important;
    font-weight: 800 !important;
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%) !important;
    box-shadow: 0 8px 18px rgba(15, 23, 42, 0.04) !important;
    transition: transform 0.18s ease, box-shadow 0.18s ease, border-color 0.18s ease !important;
}

[data-testid="stRadio"] [role="radiogroup"] label:hover {
    transform: translateY(-1px);
    box-shadow: 0 14px 24px rgba(15, 23, 42, 0.08) !important;
}

[data-testid="stRadio"] [role="radiogroup"] label[data-checked="true"] {
    background: linear-gradient(135deg, var(--primary) 0%, #6d5dfc 100%) !important;
    color: #ffffff !important;
    border-color: var(--primary) !important;
    box-shadow: 0 16px 34px rgba(79, 70, 229, 0.28) !important;
}

.stButton > button[kind="primary"] {
    height: 48px !important;
    border-radius: 14px !important;
    border: 0 !important;
    background: linear-gradient(135deg, #4f46e5 0%, #6d5dfc 48%, #8b5cf6 100%) !important;
    color: #ffffff !important;
    font-weight: 800 !important;
    box-shadow: var(--shadow-lift) !important;
    transition: transform 0.18s ease, box-shadow 0.18s ease !important;
}

.stButton > button[kind="primary"]:hover {
    background: var(--primary-dark) !important;
    transform: translateY(-2px);
    box-shadow: 0 32px 82px rgba(79, 70, 229, 0.24) !important;
}

.stButton > button,
.stDownloadButton > button {
    min-height: 40px !important;
    border-radius: 12px !important;
    border: 1px solid rgba(226, 232, 240, 0.95) !important;
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%) !important;
    color: var(--text) !important;
    font-weight: 800 !important;
    box-shadow: 0 10px 22px rgba(15, 23, 42, 0.04) !important;
}

.stButton > button:hover,
.stDownloadButton > button:hover {
    border-color: #cbd5e1 !important;
    background: linear-gradient(180deg, #ffffff 0%, #eff6ff 100%) !important;
}

[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-baseweb="select"] > div,
[data-baseweb="input"] > div,
[data-baseweb="base-input"] > div {
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%) !important;
    border-color: rgba(226, 232, 240, 0.95) !important;
    color: var(--text) !important;
    border-radius: 12px !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.7), 0 8px 18px rgba(15,23,42,0.03) !important;
}

.section-head {
    margin-bottom: 12px;
}

.section-title {
    color: var(--text);
    font-size: 1.15rem;
    font-weight: 900;
    margin: 0;
}

.section-subtitle {
    color: var(--muted);
    font-size: 0.86rem;
    font-weight: 500;
    margin-top: 4px;
}

.upload-panel-head {
    display: flex;
    align-items: flex-start;
    gap: 12px;
    margin-bottom: 12px;
}

.upload-panel-icon {
    width: 40px;
    height: 40px;
    border-radius: 12px;
    display: grid;
    place-items: center;
    background: linear-gradient(180deg, #f1f5ff 0%, #e8ecff 100%);
    color: var(--primary);
    flex-shrink: 0;
    box-shadow: 0 14px 24px rgba(79, 70, 229, 0.12);
}

.upload-panel-title {
    color: var(--text);
    font-size: 1rem;
    font-weight: 800;
}

.upload-panel-subtitle {
    color: var(--muted);
    font-size: 0.84rem;
    font-weight: 500;
    margin-top: 2px;
}

[data-testid="stFileUploaderDropzone"] {
    min-height: 176px !important;
    background:
        linear-gradient(180deg, rgba(255,255,255,0.98) 0%, rgba(246,248,253,0.96) 100%) !important;
    border: 1px dashed #cbd5e1 !important;
    border-radius: 18px !important;
    padding: 20px !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,0.9), 0 14px 28px rgba(15, 23, 42, 0.03) !important;
}

[data-testid="stFileUploaderDropzone"] section,
[data-testid="stFileUploaderDropzone"] small,
[data-testid="stFileUploaderDropzone"] span,
[data-testid="stFileUploaderDropzone"] p {
    color: var(--muted) !important;
}

[data-testid="stFileUploaderDropzone"] button {
    border-radius: 10px !important;
    border: 1px solid #dbeafe !important;
    background: #eef2ff !important;
    color: var(--primary) !important;
    font-weight: 800 !important;
}

.upload-ready-bar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 10px;
    flex-wrap: wrap;
    margin: 8px 0 10px 0;
}

.upload-ready-chip {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 12px;
    border-radius: 12px;
    background: linear-gradient(180deg, #ffffff 0%, #f2f6fc 100%);
    border: 1px solid #dbe3ef;
    color: var(--text);
    font-size: 0.84rem;
    font-weight: 700;
    box-shadow: 0 10px 20px rgba(15, 23, 42, 0.05);
}

.upload-ready-state {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 8px 10px;
    border-radius: 999px;
    background: linear-gradient(180deg, #f7fff8 0%, #ebfff0 100%);
    border: 1px solid #bbf7d0;
    color: var(--ok);
    font-size: 0.77rem;
    font-weight: 800;
    box-shadow: 0 10px 18px rgba(22, 163, 74, 0.08);
}

.upload-ready-dot {
    width: 8px;
    height: 8px;
    border-radius: 999px;
    background: currentColor;
}

.summary-grid {
    display: grid;
    grid-template-columns: repeat(6, minmax(0, 1fr));
    gap: 12px;
    margin-bottom: 10px;
}

.summary-card {
    background: linear-gradient(180deg, #ffffff 0%, #fafcff 100%);
    border: 1px solid rgba(226, 232, 240, 0.95);
    border-radius: 18px;
    padding: 16px;
    box-shadow: 0 14px 28px rgba(15, 23, 42, 0.05);
    transition: transform 0.18s ease, box-shadow 0.18s ease;
}

.summary-card:hover {
    transform: translateY(-2px);
    box-shadow: 0 20px 44px rgba(15, 23, 42, 0.08);
}

.summary-card small {
    display: block;
    color: var(--muted);
    font-size: 0.72rem;
    font-weight: 800;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    margin-bottom: 12px;
}

.summary-card b {
    display: block;
    color: var(--text);
    font-size: 1.8rem;
    line-height: 1;
    font-weight: 900;
}

.summary-card span {
    display: block;
    color: var(--muted);
    font-size: 0.8rem;
    margin-top: 10px;
    line-height: 1.45;
}

.summary-card.ok { background: linear-gradient(180deg, #f6fff8 0%, #eefcf2 100%); border-color: #bbf7d0; }
.summary-card.ok b { color: var(--ok); }
.summary-card.round { background: linear-gradient(180deg, #f7fbff 0%, #edf5ff 100%); border-color: #bfdbfe; }
.summary-card.round b { color: var(--round); }
.summary-card.div { background: linear-gradient(180deg, #fff8f8 0%, #fff1f1 100%); border-color: #fecaca; }
.summary-card.div b { color: var(--div); }
.summary-card.miss { background: linear-gradient(180deg, #fffaf5 0%, #fff4e8 100%); border-color: #fed7aa; }
.summary-card.miss b { color: var(--miss); }
.summary-card.impact { background: linear-gradient(180deg, #fffaf3 0%, #fff1de 100%); border-color: #fdba74; }
.summary-card.impact b { color: #ea580c; font-size: 1.5rem; }

.summary-note {
    margin-top: 4px;
    padding: 12px 14px;
    border-radius: 14px;
    background: #ffffff;
    border: 1px solid var(--line);
    color: #334155;
    font-size: 0.82rem;
    font-weight: 500;
    line-height: 1.6;
}

.detail-grid {
    grid-template-columns: repeat(4, minmax(0, 1fr));
    gap: 10px;
    margin: 4px 0 14px 0;
}

.detail-metric {
    background: #ffffff;
    border: 1px solid var(--line);
    border-radius: 14px;
    padding: 12px 14px;
}

.detail-metric small {
    color: var(--muted);
}

.detail-metric b {
    color: var(--text);
    font-size: 1.05rem;
}

.conference-note {
    margin-top: 0;
    background: #f8fafc;
    border: 1px solid #dbe3ef;
    color: var(--muted);
}

.table-shell {
    margin-top: 10px;
    border-radius: 16px;
    border-color: var(--line);
}

.audit-results-shell {
    max-height: 64vh;
}

.audit-table thead th {
    background: #f8fafc;
    color: var(--muted);
}

.audit-table tbody td {
    font-size: 0.86rem;
}

.export-actions {
    margin-top: 12px;
}

.export-note {
    color: var(--muted);
    font-size: 0.82rem;
    margin-bottom: 10px;
}

.compact-footer {
    margin-top: 22px;
    padding: 16px 0 6px 0;
    text-align: center;
    color: var(--muted);
    font-size: 0.78rem;
    line-height: 1.6;
}

.marketing-shell {
    display: grid;
    grid-template-columns: 1.1fr 0.9fr;
    gap: 16px;
    margin-top: 6px;
}

.marketing-card {
    position: relative;
    background: linear-gradient(135deg, rgba(255,255,255,0.98) 0%, rgba(245,247,255,0.96) 100%);
    border: 1px solid rgba(255,255,255,0.8);
    border-radius: 24px;
    padding: 24px;
    box-shadow: var(--shadow-card);
    overflow: hidden;
}

.marketing-card::before {
    content: "";
    position: absolute;
    inset: auto -70px -80px auto;
    width: 200px;
    height: 200px;
    border-radius: 999px;
    background: radial-gradient(circle, rgba(79,70,229,0.14) 0%, rgba(79,70,229,0) 68%);
    pointer-events: none;
}

.marketing-kicker {
    color: var(--primary);
    font-size: 0.78rem;
    font-weight: 800;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}

.marketing-title {
    color: var(--text);
    font-family: "Cormorant Garamond", serif;
    font-size: 2.45rem;
    line-height: 1;
    font-weight: 900;
    margin: 10px 0 12px 0;
    letter-spacing: 0.01em;
}

.marketing-copy {
    color: var(--muted);
    font-size: 0.95rem;
    line-height: 1.75;
}

.marketing-chip-row {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 16px;
}

.marketing-chip {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 10px 12px;
    border-radius: 999px;
    border: 1px solid #dbe3ef;
    background: #f8fafc;
    color: var(--text);
    font-size: 0.82rem;
    font-weight: 700;
}

.marketing-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
}

.marketing-feature {
    background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
    border: 1px solid rgba(226, 232, 240, 0.95);
    border-radius: 18px;
    padding: 16px;
    box-shadow: 0 14px 28px rgba(15, 23, 42, 0.05);
}

.marketing-feature small {
    display: block;
    color: var(--muted);
    font-size: 0.72rem;
    font-weight: 800;
    text-transform: uppercase;
    margin-bottom: 8px;
}

.marketing-feature b {
    display: block;
    color: var(--text);
    font-size: 1rem;
    margin-bottom: 8px;
}

.marketing-feature span {
    color: var(--muted);
    font-size: 0.84rem;
    line-height: 1.6;
}

.chart-card {
    margin-top: 12px;
}

.marketing-hero {
    display: grid;
    grid-template-columns: 1.05fr 0.95fr;
    gap: 18px;
    align-items: stretch;
}

.marketing-visual {
    position: relative;
    min-height: 360px;
    border-radius: 28px;
    padding: 24px;
    background:
        radial-gradient(circle at top left, rgba(79, 70, 229, 0.28), rgba(79, 70, 229, 0) 34%),
        radial-gradient(circle at bottom right, rgba(245, 158, 11, 0.22), rgba(245, 158, 11, 0) 32%),
        linear-gradient(160deg, #0f172a 0%, #172554 48%, #312e81 100%);
    box-shadow: 0 34px 90px rgba(15, 23, 42, 0.26);
    overflow: hidden;
}

.marketing-visual::before {
    content: "";
    position: absolute;
    inset: 16px;
    border-radius: 24px;
    border: 1px solid rgba(255,255,255,0.12);
    pointer-events: none;
}

.visual-glow {
    position: absolute;
    inset: auto auto -50px -30px;
    width: 220px;
    height: 220px;
    border-radius: 999px;
    background: radial-gradient(circle, rgba(255,255,255,0.18) 0%, rgba(255,255,255,0) 70%);
}

.visual-signature {
    position: relative;
    z-index: 2;
    max-width: 320px;
    filter: drop-shadow(0 18px 34px rgba(15, 23, 42, 0.35));
}

.marketing-visual .fv-wordmark-main {
    color: rgba(255,255,255,0.94);
}

.marketing-visual .fv-wordmark-main strong {
    color: #f6c56d;
}

.marketing-visual .fv-wordmark-tagline {
    color: rgba(255,255,255,0.62);
}

.visual-float-row {
    position: absolute;
    right: 22px;
    top: 24px;
    display: grid;
    gap: 12px;
    width: 190px;
    z-index: 3;
}

.visual-stat {
    padding: 14px 15px;
    border-radius: 18px;
    background: rgba(255,255,255,0.14);
    border: 1px solid rgba(255,255,255,0.18);
    backdrop-filter: blur(14px);
    box-shadow: 0 18px 42px rgba(15, 23, 42, 0.24);
}

.visual-stat small {
    display: block;
    color: rgba(255,255,255,0.72);
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    font-weight: 800;
    margin-bottom: 7px;
}

.visual-stat b {
    color: #ffffff;
    font-size: 1.18rem;
    font-weight: 900;
}

.visual-board {
    position: absolute;
    left: 24px;
    right: 24px;
    bottom: 24px;
    z-index: 2;
    padding: 18px;
    border-radius: 24px;
    background: rgba(255,255,255,0.1);
    border: 1px solid rgba(255,255,255,0.14);
    backdrop-filter: blur(16px);
    box-shadow: 0 20px 46px rgba(15, 23, 42, 0.26);
}

.visual-board-title {
    color: #ffffff;
    font-size: 0.92rem;
    font-weight: 800;
    margin-bottom: 12px;
}

.visual-bars {
    display: grid;
    gap: 10px;
}

.visual-bar-row {
    display: grid;
    grid-template-columns: 82px 1fr 38px;
    gap: 10px;
    align-items: center;
}

.visual-bar-label {
    color: rgba(255,255,255,0.78);
    font-size: 0.76rem;
    font-weight: 700;
}

.visual-bar-track {
    height: 10px;
    border-radius: 999px;
    background: rgba(255,255,255,0.14);
    overflow: hidden;
}

.visual-bar-fill {
    height: 100%;
    border-radius: inherit;
    box-shadow: 0 10px 18px rgba(0,0,0,0.18);
}

.visual-bar-value {
    color: #ffffff;
    font-size: 0.78rem;
    font-weight: 800;
    text-align: right;
}

.chart-shell .section-subtitle,
.chart-shell .summary-note,
.marketing-copy,
.marketing-feature span,
.marketing-chip,
.detail-metric small,
.summary-card span,
.summary-card small,
.compact-field-label,
.menu-caption,
.conference-note,
.export-note {
    color: #475569 !important;
}

@media (max-width: 1100px) {
    .summary-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    .marketing-shell { grid-template-columns: 1fr; }
    .marketing-hero { grid-template-columns: 1fr; }
}

@media (max-width: 760px) {
    .block-container { padding: 16px 14px 26px 14px !important; }
    .summary-grid,
    .detail-grid { grid-template-columns: 1fr; }
    .marketing-grid { grid-template-columns: 1fr; }
    .visual-float-row {
        position: static;
        width: 100%;
        margin-top: 18px;
    }
    .marketing-visual {
        min-height: 460px;
    }
    .visual-board {
        left: 18px;
        right: 18px;
        bottom: 18px;
    }
}
</style>
"""

st.markdown(COMPACT_CSS, unsafe_allow_html=True)


def icon_doc(size=20):
    return f"""
    <svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" aria-hidden="true">
        <path d="M7 3h6l4 4v14H7V3Z" stroke="currentColor" stroke-width="1.8" stroke-linejoin="round"/>
        <path d="M13 3v5h5" stroke="currentColor" stroke-width="1.8" stroke-linejoin="round"/>
        <path d="m9.5 15 2 2 4-5" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/>
    </svg>
    """


def icon_menu(size=22):
    return f"""
    <svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" aria-hidden="true">
        <path d="M4 7h16M4 12h16M4 17h16" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"/>
    </svg>
    """


def icon_search(size=18):
    return f"""
    <svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" aria-hidden="true">
        <circle cx="11" cy="11" r="7" stroke="currentColor" stroke-width="1.8"/>
        <path d="m16.5 16.5 4 4" stroke="currentColor" stroke-width="1.8" stroke-linecap="round"/>
    </svg>
    """


def fretevision_mark(size=120):
    grad_suffix = f"{size}"
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 120 120" fill="none" aria-hidden="true" xmlns="http://www.w3.org/2000/svg">'
        f'<defs><linearGradient id="fvGold{grad_suffix}" x1="60" y1="18" x2="94" y2="95" gradientUnits="userSpaceOnUse">'
        '<stop stop-color="#D6B071"/><stop offset="1" stop-color="#A6722C"/></linearGradient></defs>'
        '<text x="20" y="72" fill="#10233A" font-family="Cormorant Garamond, serif" font-size="72" font-weight="700">F</text>'
        f'<text x="57" y="88" fill="url(#fvGold{grad_suffix})" font-family="Cormorant Garamond, serif" font-size="68" font-weight="700">V</text>'
        '<path d="M18 86C41 62 67 51 104 48" stroke="#10233A" stroke-width="11" stroke-linecap="round"/>'
        '<path d="M24 88C45 69 66 59 89 55" stroke="white" stroke-width="2.8" stroke-dasharray="8 7" stroke-linecap="round"/>'
        f'<path d="M34 95C57 72 76 61 103 55" stroke="url(#fvGold{grad_suffix})" stroke-width="3" stroke-linecap="round" opacity="0.95"/></svg>'
    )


def fretevision_signature():
    return (
        '<div class="fv-signature">'
        f'<div class="fv-symbol">{fretevision_mark(190)}</div>'
        '<div class="fv-wordmark">'
        '<div class="fv-wordmark-main"><span>FRETE </span><strong>VISION</strong></div>'
        f'<div class="fv-wordmark-tagline"><span></span>{safe_text(BRAND_TAGLINE.upper())}<span></span></div>'
        '</div></div>'
    )


def br_money(value):
    missing = "Não encontrado"
    if isinstance(value, pd.Series):
        value = value.iloc[0] if not value.empty else None
    if value is None:
        return missing
    try:
        if pd.isna(value):
            return missing
    except (TypeError, ValueError):
        pass
    if value == missing:
        return value
    return format_money_br(value)

def safe_text(value):
    return html.escape(str(ui(value)))

def validate_uploaded_pdf(uploaded_file):
    if uploaded_file is None:
        return

    nome = get_upload_name(uploaded_file)
    if Path(nome).suffix.lower() != ".pdf":
        raise ValueError("Envie apenas arquivos PDF.")

    if isinstance(uploaded_file, dict):
        file_bytes = uploaded_file.get("bytes", b"")
        size = int(uploaded_file.get("size", len(file_bytes)) or 0)
    else:
        file_bytes = uploaded_file.getbuffer()
        size = int(getattr(uploaded_file, "size", len(file_bytes)) or 0)

    if not file_bytes:
        raise ValueError(f"O arquivo {nome or 'PDF'} está vazio.")

    if size > MAX_UPLOAD_SIZE_BYTES:
        limite_mb = MAX_UPLOAD_SIZE_BYTES // (1024 * 1024)
        raise ValueError(f"O arquivo {nome} excede o limite de {limite_mb} MB.")

    header = bytes(file_bytes[:PDF_HEADER_SCAN_BYTES])
    if b"%PDF-" not in header:
        raise ValueError(f"O arquivo {nome} não parece ser um PDF válido.")

def serialize_uploaded_file(uploaded_file):
    if uploaded_file is None:
        return None
    validate_uploaded_pdf(uploaded_file)
    file_bytes = bytes(uploaded_file.getbuffer())
    return {
        "name": Path(uploaded_file.name).name,
        "bytes": file_bytes,
        "size": len(file_bytes),
    }


def get_upload_name(uploaded_file):
    if uploaded_file is None:
        return ""
    if isinstance(uploaded_file, dict):
        return str(uploaded_file.get("name", ""))
    return Path(uploaded_file.name).name


def file_size(uploaded_file):
    if isinstance(uploaded_file, dict):
        size = uploaded_file.get("size", 0) or 0
    else:
        size = getattr(uploaded_file, "size", 0) or 0
    if size >= 1024 * 1024:
        return f"{size / (1024 * 1024):.1f} MB"
    return f"{size / 1024:.0f} KB"


def salvar_upload_pdf(uploaded_file, prefixo):
    if uploaded_file is None:
        return None

    validate_uploaded_pdf(uploaded_file)
    temp_dir = Path(tempfile.gettempdir()) / "fretescan_uploads"
    temp_dir.mkdir(parents=True, exist_ok=True)

    if isinstance(uploaded_file, dict):
        file_bytes = uploaded_file.get("bytes", b"")
    else:
        file_bytes = uploaded_file.getbuffer()

    prefixo_limpo = re.sub(r"[^A-Za-z0-9_-]+", "_", str(prefixo or "pdf")).strip("_") or "pdf"
    suffix = Path(get_upload_name(uploaded_file)).suffix.lower() or ".pdf"

    with tempfile.NamedTemporaryFile(
        delete=False,
        suffix=suffix,
        prefix=f"{prefixo_limpo}_",
        dir=temp_dir,
    ) as arquivo_temp:
        arquivo_temp.write(file_bytes)

    return arquivo_temp.name


def normalizar_resumo_motor(resumo):
    return {
        "total": int(resumo.get("total_analisado", 0) or 0),
        "ok": int(resumo.get("ok", 0) or 0),
        "ok_arredondamento": int(resumo.get("ok_arredondamento", 0) or 0),
        "divergentes": int(resumo.get("divergentes", 0) or 0),
        "faltantes_a": int(resumo.get("faltante_a", 0) or 0),
        "faltantes_b": int(resumo.get("faltante_b", 0) or 0),
        "dif_total_empresa": float(resumo.get("dif_total_empresa", 0) or 0),
        "dif_total_motorista": float(resumo.get("dif_total_motorista", 0) or 0),
        "impacto_absoluto": float(resumo.get("impacto_absoluto", 0) or 0),
    }


def get_result_diff_column(df):
    raw_diff_aliases = ["Maior Diferença", "Maior DiferenÃ§a", "Maior Diferen?a", "Maior Diferen??a"]
    raw_diff_col = next((col for col in raw_diff_aliases if col in df.columns), None)
    if raw_diff_col is None:
        raise KeyError("Coluna de maior diferença não encontrada no resultado da auditoria.")
    return raw_diff_col


def resolve_tolerance_value(tolerancia=None):
    if tolerancia is None:
        tolerancia = 0.50
    return float(tolerancia)


def classify_diff_bucket(value, tolerancia):
    if value is None:
        return "none"
    try:
        if pd.isna(value):
            return "none"
    except (TypeError, ValueError):
        pass
    abs_value = abs(float(value))
    if abs_value == 0:
        return "none"
    if abs_value <= resolve_tolerance_value(tolerancia):
        return "tolerance"
    return "critical"


def build_observacao(status_base, dif_empresa=None, dif_motorista=None, tolerancia=0.50):
    if status_base == "OK":
        return "Sem diferença"
    if status_base == "OK por arredondamento":
        return "Diferença dentro da tolerância"
    if status_base == "Faltante no A":
        return "Existe no GW e não existe no ATUA"
    if status_base == "Faltante no B":
        return "Existe no ATUA e não existe no GW"

    empresa_bucket = classify_diff_bucket(dif_empresa, tolerancia)
    motorista_bucket = classify_diff_bucket(dif_motorista, tolerancia)

    if empresa_bucket == "critical" and motorista_bucket == "critical":
        return "Divergência em empresa e motorista"
    if empresa_bucket == "critical":
        return "Divergência no frete empresa"
    if motorista_bucket == "critical":
        return "Divergência no frete motorista"
    return "Divergência acima da tolerância"


def format_optional_money(value):
    if value is None:
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except (TypeError, ValueError):
        pass
    if isinstance(value, (int, float, Decimal)):
        return br_money(value)
    if str(value).strip() in {"", "None", "Não encontrado"}:
        return "-"
    return ui(str(value))


def format_optional_margin(value):
    if value is None:
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except (TypeError, ValueError):
        pass
    text = ui(str(value)).strip()
    if not text or text in {"None", "Não encontrado"}:
        return "-"
    if text.endswith("%"):
        return text
    try:
        numeric = float(text.replace("%", "").replace(".", "").replace(",", "."))
        return f"{numeric:.2f}".replace(".", ",") + "%"
    except ValueError:
        return text


def format_identifier(value):
    if value is None:
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if str(value).strip() in {"", "None", "Não encontrado"}:
        return "-"
    return ui(str(value))


@st.cache_data(show_spinner=False)
def extrair_margens_gw_visual(caminho_pdf):
    caminho = str(caminho_pdf)

    if fitz is not None:
        try:
            margens = {}
            doc = fitz.open(caminho)
            try:
                for page in doc:
                    groups = defaultdict(list)
                    for x0, y0, x1, y1, text, *_ in page.get_text("blocks"):
                        groups[round(x0, 1)].append(text)

                    for texts in groups.values():
                        lines = []
                        for text in texts:
                            lines.extend([line.strip() for line in text.splitlines() if line.strip()])

                        ctes = [line for line in lines if re.fullmatch(r"0*\d{4,}", line)]
                        pcts = [line for line in lines if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})*,\d{2}%", line)]
                        if ctes and pcts:
                            margens[str(int(ctes[-1]))] = pcts[-1]
            finally:
                doc.close()

            if margens:
                return margens
        except Exception:
            pass

    try:
        margens = {}
        with pdfplumber.open(caminho) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    match = RE_GW_MARGIN_VISUAL.match(line.strip())
                    if not match:
                        continue
                    cte = str(int(match.group(1)))
                    margens[cte] = match.group(2)
        return margens
    except Exception:
        return {}


def aplicar_margens_gw_visual(df, caminho_gw):
    if df is None or df.empty or "Margem B" not in df.columns or not caminho_gw:
        return df
    margem_map = extrair_margens_gw_visual(caminho_gw)
    if not margem_map:
        return df

    enriched = df.copy()
    current_values = enriched["Margem B"].tolist()
    ctes = [format_identifier(value).lstrip("0") for value in enriched["CTE"].tolist()]

    enriched["Margem B"] = [
        current if pd.notna(current) and str(current).strip() not in ["", "None"] else margem_map.get(cte, None)
        for current, cte in zip(current_values, ctes)
    ]
    return enriched


def resolve_gw_visual_source_path():
    caminho = st.session_state.get("caminho_b_temp")
    if caminho and Path(str(caminho)).exists():
        return str(caminho)

    stored = st.session_state.get("up_b_stored")
    if stored is None:
        return None

    try:
        caminho = salvar_upload_pdf(stored, "GW")
    except Exception:
        return None
    return caminho if caminho and Path(str(caminho)).exists() else None


def ensure_gw_margin_visual(df, caminho_gw=None):
    if df is None or df.empty:
        return df

    margin_column = None
    if "Margem B" in df.columns:
        margin_column = "Margem B"
    elif "Margem GW" in df.columns:
        margin_column = "Margem GW"

    if not margin_column:
        return df

    existing = df[margin_column].fillna("").astype(str).str.strip()
    if existing[~existing.isin(["", "-", "None", "Não encontrado", "nan"])].any():
        return df

    resolved_path = caminho_gw or resolve_gw_visual_source_path()
    if not resolved_path:
        return df

    if margin_column == "Margem B":
        return aplicar_margens_gw_visual(df, resolved_path)

    enriched = aplicar_margens_gw_visual(df.rename(columns={"Margem GW": "Margem B"}), resolved_path)
    return enriched.rename(columns={"Margem B": "Margem GW"})


def clear_export_caches():
    for fn in [build_export_summary_rows, build_excel_bytes, build_executive_pdf_bytes, build_detailed_pdf_bytes]:
        try:
            fn.clear()
        except Exception:
            pass


def prepare_conference_dataframe(df, tolerancia=0.50):
    df = ensure_gw_margin_visual(df)
    diff_col = get_result_diff_column(df)
    prepared = df.copy().rename(
        columns={
            diff_col: "Maior Diferença",
            "Dif Empresa": "Dif. Empresa",
            "Dif Motorista": "Dif. Motorista",
            "Margem B": "Margem GW",
        }
    )
    prepared = prepared.loc[:, ~prepared.columns.duplicated()].copy()

    if "Margem GW" not in prepared.columns:
        prepared["Margem GW"] = None

    tolerancia_valor = resolve_tolerance_value(tolerancia)
    prepared["_Status Base"] = prepared["Status"].astype(str)
    prepared["_Dif Empresa Num"] = pd.to_numeric(prepared["Dif. Empresa"], errors="coerce")
    prepared["_Dif Motorista Num"] = pd.to_numeric(prepared["Dif. Motorista"], errors="coerce")
    prepared["_CTE Num"] = pd.to_numeric(prepared["CTE"], errors="coerce")
    prepared["_Maior Diferença Num"] = pd.to_numeric(prepared["Maior Diferença"], errors="coerce").fillna(0.0)
    prepared["_Status Ordem"] = prepared["_Status Base"].map(STATUS_ORDER_MAP).fillna(99)
    prepared["_Empresa Visual"] = prepared["_Dif Empresa Num"].map(lambda value: classify_diff_bucket(value, tolerancia_valor))
    prepared["_Motorista Visual"] = prepared["_Dif Motorista Num"].map(lambda value: classify_diff_bucket(value, tolerancia_valor))
    prepared["_Linha Visual"] = prepared["_Status Base"].map(
        lambda status: "missing" if status in ["Faltante no A", "Faltante no B"] else "normal"
    )
    prepared["Observação"] = prepared.apply(
        lambda row: build_observacao(
            row["_Status Base"],
            row["_Dif Empresa Num"],
            row["_Dif Motorista Num"],
            tolerancia_valor,
        ),
        axis=1,
    )
    prepared["Status"] = prepared["_Status Base"].replace(STATUS_DISPLAY_MAP)
    return prepared


def apply_conference_filters(prepared, scope_label, min_diff, search_text, order_label):
    visible = prepared.copy()

    if search_text:
        visible = visible[visible["CTE"].astype(str).str.contains(search_text.strip(), case=False, na=False)]

    if scope_label == "Críticos":
        visible = visible[visible["_Status Base"].isin(["Divergente", "Faltante no A", "Faltante no B"])]
    elif scope_label == "Divergentes reais":
        visible = visible[visible["_Status Base"] == "Divergente"]
    elif scope_label == "Diferenças dentro da tolerância":
        visible = visible[visible["_Status Base"] == "OK por arredondamento"]
    elif scope_label == "Faltantes":
        visible = visible[visible["_Status Base"].isin(["Faltante no A", "Faltante no B"])]
    elif scope_label == "OK sem diferença":
        visible = visible[visible["_Status Base"] == "OK"]

    visible = visible[visible["_Maior Diferença Num"] >= float(min_diff or 0.0)]

    if order_label == "Maior diferença":
        visible = visible.sort_values(["_Maior Diferença Num", "_CTE Num"], ascending=[False, True], na_position="last")
    elif order_label == "CTE crescente":
        visible = visible.sort_values(["_CTE Num", "CTE"], na_position="last")
    else:
        visible = visible.sort_values(["_Status Ordem", "_Maior Diferença Num", "_CTE Num"], ascending=[True, False, True], na_position="last")

    return visible


def build_conference_display_table(prepared):
    table_columns = [
        "CTE",
        "Status",
        "Empresa A",
        "Empresa B",
        "Dif. Empresa",
        "Motorista A",
        "Motorista B",
        "Dif. Motorista",
        "Maior Diferença",
        "Margem GW",
        "Observação",
    ]
    money_columns = [
        "Empresa A",
        "Empresa B",
        "Dif. Empresa",
        "Motorista A",
        "Motorista B",
        "Diferença",
        "Maior Diferença",
    ]

    visible = prepared[[column for column in table_columns if column in prepared.columns]].copy()
    visible = visible.fillna("-").replace({None: "-", "": "-", "None": "-", "Não encontrado": "-"})
    visible["CTE"] = visible["CTE"].map(format_identifier)

    for column in ["Empresa A", "Empresa B", "Dif. Empresa", "Motorista A", "Motorista B", "Dif. Motorista", "Maior Diferença", "Margem GW"]:
        if column in visible.columns:
            if column == "Margem GW":
                visible[column] = visible[column].map(format_optional_margin)
            else:
                visible[column] = visible[column].map(format_optional_money)

    visible = visible.rename(columns={"Dif. Motorista": "Diferença"})
    visible["Observação"] = visible["Observação"].map(ui)
    return visible, money_columns


def build_detailed_counts_html(resumo):
    return (
        '<div class="detail-grid">'
        f'<div class="detail-metric"><small>Divergentes reais</small><b>{safe_text(resumo["divergentes"])}</b></div>'
        f'<div class="detail-metric"><small>OK por arredondamento</small><b>{safe_text(resumo["ok_arredondamento"])}</b></div>'
        f'<div class="detail-metric"><small>Faltante no A</small><b>{safe_text(resumo["faltantes_a"])}</b></div>'
        f'<div class="detail-metric"><small>Faltante no B</small><b>{safe_text(resumo["faltantes_b"])}</b></div>'
        '</div>'
        '<div class="conference-note">Diferenças dentro da tolerância são exibidas para conferência, mas não compõem o impacto crítico.</div>'
    )


@st.cache_data(show_spinner=False)
def build_export_summary_rows(df, resumo, tolerancia=0.50):
    prepared = prepare_conference_dataframe(df, tolerancia)
    criticos = prepared[prepared["_Status Base"].isin(["Divergente", "Faltante no A", "Faltante no B"])].copy()
    dif_empresa_critica = float(round(pd.to_numeric(criticos.get("Dif. Empresa"), errors="coerce").fillna(0.0).sum(), 2)) if "Dif. Empresa" in criticos.columns else 0.0
    dif_motorista_critica = float(round(pd.to_numeric(criticos.get("Dif. Motorista"), errors="coerce").fillna(0.0).sum(), 2)) if "Dif. Motorista" in criticos.columns else 0.0
    return [
        ["Total analisado", str(resumo["total"])],
        ["OK", str(resumo["ok"])],
        ["OK por arredondamento", str(resumo["ok_arredondamento"])],
        ["Divergentes reais", str(resumo["divergentes"])],
        ["Faltantes no A", str(resumo["faltantes_a"])],
        ["Faltantes no B", str(resumo["faltantes_b"])],
        ["Diferença Empresa — Crítico Total", br_money(dif_empresa_critica)],
        ["Diferença Motorista — Crítico Total", br_money(dif_motorista_critica)],
        ["Impacto Crítico Total", br_money(resumo["impacto_absoluto"])],
    ]


def build_export_criteria_rows(nome_a, nome_b, tolerancia):
    return [
        ["Critério", "Descrição"],
        ["Arquivos comparados", f"{nome_a} x {nome_b}"],
        ["Tolerância configurada", br_money(tolerancia)],
        ["Dif. Empresa", "Empresa A - Empresa B"],
        ["Diferença", "Motorista A - Motorista B"],
        ["Regra visual", "Diferenças dentro da tolerância aparecem para conferência, mas não entram no impacto crítico."],
    ]


def build_export_frames(df, tolerancia=0.50):
    prepared = prepare_conference_dataframe(df, tolerancia)
    display_df, _ = build_conference_display_table(prepared)
    divergentes_df = display_df[display_df["Status"] == "Divergente"].copy()
    tolerancia_df = display_df[display_df["Status"] == "OK Arred."].copy()
    faltantes_df = display_df[display_df["Status"].isin(["Faltante no A", "Faltante no B"])].copy()
    criticos_df = display_df[display_df["Status"].isin(["Divergente", "Faltante no A", "Faltante no B"])].copy()
    return prepared, display_df, divergentes_df, tolerancia_df, faltantes_df, criticos_df


@st.cache_data(show_spinner=False)
def build_excel_bytes(df, resumo, nome_a, nome_b, tolerancia):
    import io
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    prepared, display_df, divergentes_df, _, faltantes_df, criticos_df = build_export_frames(df, tolerancia)
    summary_rows = build_export_summary_rows(df, resumo, tolerancia)
    criteria_rows = build_export_criteria_rows(nome_a, nome_b, tolerancia)

    header_fill = PatternFill(fill_type="solid", start_color="10233A", end_color="10233A")
    accent_fill = PatternFill(fill_type="solid", start_color="B6843C", end_color="B6843C")
    even_fill = PatternFill(fill_type="solid", start_color="F8FAFC", end_color="F8FAFC")
    missing_fill = PatternFill(fill_type="solid", start_color="FFF7ED", end_color="FFF7ED")
    diff_critical_fill = PatternFill(fill_type="solid", start_color="FEE2E2", end_color="FEE2E2")
    diff_tolerance_fill = PatternFill(fill_type="solid", start_color="E8F3FF", end_color="E8F3FF")
    border = Border(
        left=Side(style="thin", color="D5DCE6"),
        right=Side(style="thin", color="D5DCE6"),
        top=Side(style="thin", color="D5DCE6"),
        bottom=Side(style="thin", color="D5DCE6"),
    )
    status_fills = {
        "OK": PatternFill(fill_type="solid", start_color="DCFCE7", end_color="DCFCE7"),
        "OK Arred.": PatternFill(fill_type="solid", start_color="DBEAFE", end_color="DBEAFE"),
        "Divergente": PatternFill(fill_type="solid", start_color="FEE2E2", end_color="FEE2E2"),
        "Faltante no A": PatternFill(fill_type="solid", start_color="FEF3C7", end_color="FEF3C7"),
        "Faltante no B": PatternFill(fill_type="solid", start_color="FEF3C7", end_color="FEF3C7"),
    }
    company_columns = {"Empresa A", "Empresa B", "Dif. Empresa"}
    driver_columns = {"Motorista A", "Motorista B", "Diferença"}

    def visual_fill(bucket):
        if bucket == "critical":
            return diff_critical_fill
        if bucket == "tolerance":
            return diff_tolerance_fill
        return None

    def style_header(ws, row_index, accent=False):
        fill = accent_fill if accent else header_fill
        for cell in ws[row_index]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def autosize_columns(ws):
        for column_cells in ws.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width + 3, 42)

    def write_table_sheet(ws, title, frame, source_prepared):
        ws.append([title])
        ws["A1"].font = Font(size=14, bold=True, color="10233A")
        ws.append([f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        ws.append([f"Arquivos: {nome_a} x {nome_b}"])
        ws.append([f"Tolerância configurada: {br_money(tolerancia)}"])
        ws.append(["Dif. Empresa = Empresa A - Empresa B"])
        ws.append(["Diferença = Motorista A - Motorista B"])
        ws.append(["Diferenças dentro da tolerância aparecem para conferência, mas não entram no impacto crítico."])
        ws.append([])
        ws.append(list(frame.columns))
        style_header(ws, ws.max_row)
        ws.freeze_panes = "A9"

        for excel_row, (source_index, row) in enumerate(frame.iterrows(), start=10):
            ws.append(row.tolist())
            meta = source_prepared.loc[source_index] if source_index in source_prepared.index else None
            is_missing = bool(meta is not None and meta.get("_Linha Visual") == "missing")
            empresa_fill = visual_fill(meta.get("_Empresa Visual")) if meta is not None else None
            motorista_fill = visual_fill(meta.get("_Motorista Visual")) if meta is not None else None

            for col_pos, cell in enumerate(ws[excel_row], start=1):
                column_name = frame.columns[col_pos - 1]
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if is_missing:
                    cell.fill = missing_fill
                elif column_name == "Status":
                    cell.fill = status_fills.get(row.get("Status"), even_fill if excel_row % 2 == 0 else PatternFill(fill_type=None))
                elif column_name in company_columns and empresa_fill:
                    cell.fill = empresa_fill
                elif column_name in driver_columns and motorista_fill:
                    cell.fill = motorista_fill
                elif excel_row % 2 == 0:
                    cell.fill = even_fill

    buf = io.BytesIO()
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumo Executivo"
    ws_summary.append(["Frete Vision — Relatório de Auditoria Logística"])
    ws_summary["A1"].font = Font(size=16, bold=True, color="10233A")
    ws_summary.append([f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws_summary.append([f"Arquivo A: {nome_a}"])
    ws_summary.append([f"Arquivo B: {nome_b}"])
    ws_summary.append([f"Tolerância configurada: {br_money(tolerancia)}"])
    ws_summary.append([])
    ws_summary.append(["Métrica", "Valor"])
    style_header(ws_summary, ws_summary.max_row)
    for label, value in summary_rows:
        ws_summary.append([label, value])
        for cell in ws_summary[ws_summary.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="top")
    ws_summary.append([])
    ws_summary.append(["Critério", "Descrição"])
    style_header(ws_summary, ws_summary.max_row, accent=True)
    for label, value in criteria_rows[1:]:
        ws_summary.append([label, value])
        for cell in ws_summary[ws_summary.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    write_table_sheet(wb.create_sheet("Críticos e Faltantes"), "Divergentes e Faltantes", criticos_df, prepared)
    write_table_sheet(wb.create_sheet("Conferência Detalhada"), "Conferência Detalhada", display_df, prepared)
    write_table_sheet(wb.create_sheet("Divergentes Reais"), "Divergentes Reais", divergentes_df, prepared)
    write_table_sheet(wb.create_sheet("Faltantes"), "Faltantes", faltantes_df, prepared)

    for ws in wb.worksheets:
        autosize_columns(ws)

    wb.save(buf)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def build_executive_pdf_bytes(df, resumo, nome_a, nome_b, tolerancia):
    import io
    from datetime import datetime

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _, _, divergentes_df, _, faltantes_df, criticos_df = build_export_frames(df, tolerancia)
    summary_rows = [["Métrica", "Valor"]] + build_export_summary_rows(df, resumo, tolerancia)
    criteria_rows = build_export_criteria_rows(nome_a, nome_b, tolerancia)

    def append_table(story, title, frame, columns, widths, empty_message, heading_style, body_style, max_rows=30, new_page=False):
        if new_page:
            story.append(PageBreak())
        story.append(Paragraph(title, heading_style))
        story.append(Spacer(1, 0.18 * cm))
        if frame.empty:
            story.append(Paragraph(empty_message, body_style))
            story.append(Spacer(1, 0.35 * cm))
            return

        for start in range(0, len(frame), max_rows):
            chunk = frame.iloc[start:start + max_rows]
            data = [columns] + [[str(row[col]) for col in columns] for _, row in chunk.iterrows()]
            table = Table(data, colWidths=widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("LEADING", (0, 0), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DCE6")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.22 * cm))
            if start + max_rows < len(frame):
                story.append(PageBreak())
                story.append(Paragraph(f"{title} (continuação)", heading_style))
                story.append(Spacer(1, 0.18 * cm))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    body_style = styles["BodyText"]
    note_style = ParagraphStyle("ExecutivePdfNote", parent=styles["BodyText"], fontSize=9, leading=12, textColor=colors.HexColor("#4B5563"))

    story = [
        Paragraph("Frete Vision — Relatório de Auditoria Logística", title_style),
        Spacer(1, 0.2 * cm),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", body_style),
        Paragraph(f"Arquivo A: {safe_text(nome_a)} | Arquivo B: {safe_text(nome_b)}", body_style),
        Paragraph(f"Tolerância configurada: {br_money(tolerancia)}", body_style),
        Spacer(1, 0.18 * cm),
        Paragraph("Dif. Empresa = Empresa A - Empresa B", note_style),
        Paragraph("Diferença = Motorista A - Motorista B", note_style),
        Paragraph("Diferenças dentro da tolerância aparecem para conferência, mas não entram no impacto crítico.", note_style),
        Spacer(1, 0.35 * cm),
        Paragraph("Resumo geral", heading_style),
        Spacer(1, 0.18 * cm),
    ]

    summary_table = Table(summary_rows, colWidths=[8.4 * cm, 5.0 * cm], repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5DCE6")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * cm))

    criteria_table = Table(criteria_rows, colWidths=[5.8 * cm, 16.2 * cm], repeatRows=1)
    criteria_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B6843C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5DCE6")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFAF6")]),
    ]))
    story.append(criteria_table)
    story.append(Spacer(1, 0.32 * cm))

    common_columns = ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa", "Motorista A", "Motorista B", "Diferença", "Maior Diferença", "Margem GW", "Observação"]
    common_widths = [1.25 * cm, 2.0 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.9 * cm, 1.7 * cm, 5.35 * cm]

    append_table(
        story,
        "Principais divergências e faltantes",
        criticos_df,
        common_columns,
        common_widths,
        "Nenhum CTE crítico encontrado.",
        heading_style,
        body_style,
        max_rows=26,
    )
    append_table(
        story,
        "Divergentes reais",
        divergentes_df,
        common_columns,
        common_widths,
        "Nenhuma divergência real encontrada.",
        heading_style,
        body_style,
        max_rows=26,
        new_page=True,
    )
    append_table(
        story,
        "Faltantes",
        faltantes_df,
        common_columns,
        common_widths,
        "Nenhum faltante encontrado.",
        heading_style,
        body_style,
        max_rows=26,
        new_page=True,
    )

    doc.build(story)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def build_detailed_pdf_bytes(df, resumo, nome_a, nome_b, tolerancia):
    import io
    from datetime import datetime

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    _, display_df, _, tolerancia_df, faltantes, criticos = build_export_frames(df, tolerancia)
    summary_rows = [["Métrica", "Valor"]] + build_export_summary_rows(df, resumo, tolerancia)
    criteria_rows = build_export_criteria_rows(nome_a, nome_b, tolerancia)

    def append_table(story, title, frame, columns, widths, empty_message, heading_style, body_style, max_rows=45, new_page=False):
        if new_page:
            story.append(PageBreak())
        story.append(Paragraph(title, heading_style))
        story.append(Spacer(1, 0.18 * cm))
        if frame.empty:
            story.append(Paragraph(empty_message, body_style))
            story.append(Spacer(1, 0.35 * cm))
            return

        for start in range(0, len(frame), max_rows):
            chunk = frame.iloc[start:start + max_rows]
            data = [columns] + [[str(row[col]) for col in columns] for _, row in chunk.iterrows()]
            table = Table(data, colWidths=widths, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("LEADING", (0, 0), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DCE6")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.22 * cm))
            if start + max_rows < len(frame):
                story.append(PageBreak())
                story.append(Paragraph(f"{title} (continuação)", heading_style))
                story.append(Spacer(1, 0.18 * cm))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    body_style = styles["BodyText"]
    note_style = ParagraphStyle("PdfNote", parent=styles["BodyText"], fontSize=9, leading=12, textColor=colors.HexColor("#4B5563"))

    story = [
        Paragraph("Frete Vision — Relatório de Auditoria Logística", title_style),
        Spacer(1, 0.2 * cm),
        Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", body_style),
        Paragraph(f"Arquivo A: {safe_text(nome_a)} | Arquivo B: {safe_text(nome_b)}", body_style),
        Paragraph(f"Tolerância configurada: {br_money(tolerancia)}", body_style),
        Spacer(1, 0.18 * cm),
        Paragraph("Dif. Empresa = Empresa A - Empresa B", note_style),
        Paragraph("Diferença = Motorista A - Motorista B", note_style),
        Paragraph("Diferenças dentro da tolerância aparecem para conferência, mas não entram no impacto crítico.", note_style),
        Spacer(1, 0.35 * cm),
        Paragraph("Resumo executivo", heading_style),
        Spacer(1, 0.18 * cm),
    ]

    summary_table = Table(summary_rows, colWidths=[7.0 * cm, 4.4 * cm], repeatRows=1)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10233A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5DCE6")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("Critérios usados", heading_style))
    story.append(Spacer(1, 0.18 * cm))
    criteria_table = Table(criteria_rows, colWidths=[6.0 * cm, 14.0 * cm], repeatRows=1)
    criteria_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B6843C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5DCE6")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFAF6")]),
    ]))
    story.append(criteria_table)
    story.append(Spacer(1, 0.28 * cm))
    story.append(Paragraph("Diferenças dentro da tolerância aparecem para conferência, mas não entram no impacto crítico.", note_style))
    story.append(Spacer(1, 0.35 * cm))

    append_table(
        story,
        "Lista de CTEs críticos",
        criticos,
        ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa", "Motorista A", "Motorista B", "Diferença", "Maior Diferença", "Margem GW", "Observação"],
        [1.25 * cm, 2.0 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.9 * cm, 1.7 * cm, 5.35 * cm],
        "Nenhum CTE crítico encontrado.",
        heading_style,
        body_style,
        max_rows=26,
    )
    append_table(
        story,
        "Lista de diferenças dentro da tolerância",
        tolerancia_df,
        ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa", "Motorista A", "Motorista B", "Diferença", "Maior Diferença", "Margem GW", "Observação"],
        [1.25 * cm, 2.0 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.9 * cm, 1.7 * cm, 5.35 * cm],
        "Nenhuma diferença dentro da tolerância encontrada.",
        heading_style,
        body_style,
        max_rows=26,
        new_page=True,
    )
    append_table(
        story,
        "Lista de faltantes",
        faltantes,
        ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa", "Motorista A", "Motorista B", "Diferença", "Maior Diferença", "Margem GW", "Observação"],
        [1.25 * cm, 2.0 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.9 * cm, 1.7 * cm, 5.35 * cm],
        "Nenhum faltante encontrado.",
        heading_style,
        body_style,
        max_rows=26,
        new_page=True,
    )
    append_table(
        story,
        "Tabela completa para conferência",
        display_df,
        ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa", "Motorista A", "Motorista B", "Diferença", "Maior Diferença", "Margem GW", "Observação"],
        [1.25 * cm, 2.0 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm, 1.9 * cm, 1.7 * cm, 5.35 * cm],
        "Nenhum registro disponível para conferência.",
        heading_style,
        body_style,
        max_rows=26,
        new_page=True,
    )

    doc.build(story)
    return buf.getvalue()


def build_file_debug(caminho):
    if not caminho:
        return {"caminho": None, "existe": False, "tamanho_bytes": 0}

    arquivo = Path(caminho)
    existe = arquivo.exists()
    tamanho = arquivo.stat().st_size if existe else 0
    return {"caminho": str(arquivo), "existe": existe, "tamanho_bytes": tamanho}


def build_debug_preview(registros):
    registros = registros or {}
    top_ctes = sorted(registros.keys(), key=lambda x: int(x))[:10]

    return {
        "quantidade_ctes": int(len(registros)),
        "ctes": top_ctes,
        "empresa": [format_money_br(registros[cte]["empresa"]) for cte in top_ctes],
        "motorista": [format_money_br(registros[cte]["motorista"]) for cte in top_ctes],
        "paginas": [registros[cte].get("pagina") for cte in top_ctes],
    }


def reset_audit_state():
    for key in ["df_res", "resumo", "nome_a", "nome_b", "tol", "audit_debug"]:
        st.session_state.pop(key, None)


def set_audit_error(message, details=None):
    st.session_state.audit_error = ui(message)
    st.session_state.audit_error_details = [ui(detail) for detail in (details or [])]
    st.session_state.processing = False
    st.rerun()


def titulo_erro_processamento(mensagem):
    texto = (mensagem or "").upper()
    if "ATUA" in texto and "GW" in texto:
        return "Falha na leitura do ATUA e do GW."
    if "ATUA" in texto:
        return "Falha na leitura do ATUA."
    if "GW" in texto:
        return "Falha na leitura do GW."
    return "Não foi possível processar os PDFs enviados."


def set_audit_debug(registros_a=None, registros_b=None, warn_a=None, warn_b=None, caminho_a=None, caminho_b=None, timings=None):
    st.session_state.audit_debug = {
        "timings": {
            label: round(float(seconds), 3)
            for label, seconds in (timings or {}).items()
        },
        "ATUA": {
            **build_debug_preview(registros_a),
            "warnings": [ui(item) for item in (warn_a or [])],
            "arquivo": build_file_debug(caminho_a),
        },
        "GW": {
            **build_debug_preview(registros_b),
            "warnings": [ui(item) for item in (warn_b or [])],
            "arquivo": build_file_debug(caminho_b),
        },
    }


def render_audit_debug():
    debug = st.session_state.get("audit_debug")
    if not debug:
        return

    expanded = bool(st.session_state.get("audit_error"))
    with st.expander("Debug da leitura", expanded=expanded):
        timings = debug.get("timings", {})
        if timings:
            st.markdown("**Tempo por etapa**")
            for label, seconds in timings.items():
                st.write(f"{label}: {seconds:.2f}s")
            st.divider()

        col_a, col_b = st.columns(2, gap="large")
        blocks = [
            (col_a, "ATUA", "Empresa A", "Motorista A"),
            (col_b, "GW", "Empresa B", "Motorista B"),
        ]
        for column, label, emp_label, mot_label in blocks:
            data = debug.get(label, {})
            arquivo = data.get("arquivo", {})
            with column:
                st.markdown(f"**{label}**")
                st.write(f"Arquivo salvo: {arquivo.get('caminho') or '-'}")
                st.write(f"Arquivo existe: {'Sim' if arquivo.get('existe') else 'Não'}")
                st.write(f"Tamanho em bytes: {arquivo.get('tamanho_bytes', 0)}")
                st.write(f"Quantidade de CTEs encontrados: {data.get('quantidade_ctes', 0)}")
                st.write(f"Primeiros 10 CTEs: {data.get('ctes', [])}")
                st.write(f"Primeiros 10 valores {emp_label}: {data.get('empresa', [])}")
                st.write(f"Primeiros 10 valores {mot_label}: {data.get('motorista', [])}")
                st.write(f"Primeiras páginas lidas: {data.get('paginas', [])}")
                for warning in data.get("warnings", []):
                    st.warning(warning)

def render_sidebar():
    with st.sidebar:
        st.markdown(
            f"""
            <div class="sidebar-brand">
                <div class="logo-mark">{fretevision_mark(54)}</div>
                <div>
                    <div class="brand-name">{safe_text(BRAND_NAME)}</div>
                    <div class="brand-caption">{safe_text(BRAND_TAGLINE)}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        page = st.radio(
            "Menu",
            ["Dashboard", "Nova Auditoria", "Histórico", "Relatórios", "Configurações", "Sobre"],
            index=1,
            label_visibility="collapsed",
        )
        st.markdown(
            """
            <div class="sidebar-profile">
                <div class="user-chip">
                    <div class="avatar">FV</div>
                    <div>
                        <div class="user-name">FRETE VISION</div>
                        <div class="user-email">Plataforma de inteligência logística</div>
                    </div>
                </div>
                <div class="progress-mini">
                    <div class="progress-label"><span>Status do sistema</span><span>Online</span></div>
                    <div class="progress-bar"><div class="progress-fill"></div></div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    return page
def render_topbar(title):
    st.markdown(
        f"""
        <div class="topbar">
            <div class="topbar-left">
                <div class="hamburger">{icon_menu()}</div>
                <div class="page-name">{safe_text(title)}</div>
            </div>
            <div class="topbar-right">
                <div class="search-box">{icon_doc(16)}<span>{safe_text(BRAND_TAGLINE)}</span></div>
                <div class="top-icon">&#9679;</div>
                <div class="top-icon">&#9670;<span class="dot">3</span></div>
                <div class="top-icon">&#9684;<span class="dot">12</span></div>
                <div class="top-icon">&#9635;<span class="dot">1</span></div>
                <div class="top-avatar">FV</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
def render_breadcrumb(section, current):
    if str(section).strip().lower() == "fretescan":
        section = BRAND_NAME
    st.markdown(
        f"""<div class="breadcrumb"><b>{safe_text(section)}</b>&nbsp;/<span>{safe_text(current)}</span></div>""",
        unsafe_allow_html=True,
    )


def render_hero():
    st.markdown(
        f"""
        <div class="panel span-12 hero-shell">
            <div class="panel-body">
                <div class="hero-layout">
                    <div class="hero-copy">
                        <div class="hero-kicker">{safe_text(BRAND_PLATFORM)}</div>
                        <h1>{safe_text(BRAND_TAGLINE)}.</h1>
                        <p>Auditoria avançada, cálculo preciso de frete, viabilidade de rotas e controle operacional para operações que precisam proteger margem, acelerar decisões e enxergar a logística com clareza executiva.</p>
                        <div class="hero-actions">
                            <span class="hero-action primary">Auditoria documental com leitura rápida</span>
                            <span class="hero-action">Conferência por CTE com tolerância financeira</span>
                        </div>
                        <div class="hero-stats">
                            <div class="hero-stat"><b>23%</b><span>de redução média em perdas invisíveis na conferência.</span></div>
                            <div class="hero-stat"><b>100%</b><span>de rastreabilidade no cruzamento dos relatórios.</span></div>
                            <div class="hero-stat"><b>+50k</b><span>CTEs prontos para leitura executiva em escala.</span></div>
                            <div class="hero-stat"><b>+200</b><span>operações com visão mais segura de resultado.</span></div>
                        </div>
                    </div>
                    <div class="hero-visual">
                        <div class="hero-brand-card">
                            {fretevision_signature()}
                            <div class="hero-proof-grid">
                                <div class="hero-proof"><small>Solução</small><b>Auditoria de Faturas</b><span>Comparação objetiva entre os dois lados da operação.</span></div>
                                <div class="hero-proof"><small>Cálculo</small><b>Frete 360º</b><span>Leitura financeira pronta para decisão executiva.</span></div>
                                <div class="hero-proof"><small>Rotas</small><b>Viabilidade</b><span>Mais contexto para validar cenário e margem.</span></div>
                                <div class="hero-proof"><small>Operação</small><b>Controle</b><span>Status, histórico e governança no mesmo fluxo.</span></div>
                            </div>
                        </div>
                    </div>
                </div>
                <div class="feature-strip">
                    <div class="feature-card"><small>Frente 01</small><b>Auditoria de Faturas</b><span>Identifique divergências, faltantes e arredondamentos com leitura clara por CTE.</span></div>
                    <div class="feature-card"><small>Frente 02</small><b>Cálculo de Frete 360º</b><span>Transforme números soltos em visão financeira utilizável pelo time gestor.</span></div>
                    <div class="feature-card"><small>Frente 03</small><b>Viabilidade de Rotas</b><span>Compare cenários com mais inteligência antes de aceitar perda de margem.</span></div>
                    <div class="feature-card"><small>Frente 04</small><b>Controle Operacional</b><span>Consolide evidência documental e histórico para sustentar a tomada de decisão.</span></div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
def render_upload_box(label, key):
    stored_key = f"{key}_stored"
    version_key = f"{key}_widget_version"
    legacy_upload = st.session_state.get(key)
    if st.session_state.get(stored_key) is None and legacy_upload is not None and hasattr(legacy_upload, "getvalue"):
        st.session_state[stored_key] = serialize_uploaded_file(legacy_upload)
        st.session_state.pop(key, None)

    current_file = st.session_state.get(stored_key)
    is_loaded = current_file is not None
    badge = '<span class="upload-badge loaded">Carregado</span>' if is_loaded else '<span class="upload-badge">Aguardando</span>'
    widget_version = int(st.session_state.get(version_key, 0) or 0)
    widget_key = f"{key}_widget_{widget_version}"

    with st.container(border=True):
        st.markdown(
            f"""
            <div class="upload-card{' is-loaded' if is_loaded else ''}">
                <div class="upload-shell">
                    <div class="upload-head">
                        <div class="small-icon">{icon_doc(18)}</div>
                        <div class="upload-copy">
                            <div class="upload-title-row">
                                <div class="upload-name">{safe_text(label)}</div>
                                {badge}
                            </div>
                            <div class="upload-help">Arraste ou selecione o PDF do período.</div>
                        </div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if not is_loaded:
            uploaded = st.file_uploader(label, type=["pdf"], key=widget_key, label_visibility="collapsed")
            if uploaded:
                st.session_state[stored_key] = serialize_uploaded_file(uploaded)
                st.session_state.pop(widget_key, None)
                st.rerun()
            return current_file

        st.markdown(
            f"""
            <div class="upload-inline-status">
                <span class="upload-inline-dot"></span>
                <strong>Arquivo pronto</strong>
                <span>{safe_text(get_upload_name(current_file))} - {file_size(current_file)}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        action_col_a, action_col_b = st.columns(2)
        with action_col_a:
            if st.button("Trocar arquivo", key=f"{key}_replace", use_container_width=True):
                st.session_state.pop(stored_key, None)
                st.session_state.pop(widget_key, None)
                st.session_state[version_key] = widget_version + 1
                reset_audit_state()
                st.session_state.audit_error = None
                st.session_state.audit_error_details = []
                st.rerun()
        with action_col_b:
            if st.button("Remover arquivo", key=f"{key}_remove", use_container_width=True):
                st.session_state.pop(stored_key, None)
                st.session_state.pop(widget_key, None)
                st.session_state[version_key] = widget_version + 1
                reset_audit_state()
                st.session_state.audit_error = None
                st.session_state.audit_error_details = []
                st.rerun()
        return current_file
def render_upload_section():
    with st.container(border=True):
        st.markdown(
            """
            <div class="panel-title" style="min-height:64px;padding:0;border-bottom:1px solid #eeedf3;margin-bottom:18px;">
                <div>
                    <h3>Enviar documentos</h3>
                    <div class="panel-subtitle">Selecione os dois relatórios do mesmo período para iniciar a auditoria.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        col_a, col_b = st.columns(2, gap="medium")
        with col_a:
            file_a = render_upload_box("Relatório A - ATUA", "up_a")
        with col_b:
            file_b = render_upload_box("Relatório B - GW", "up_b")
    return file_a, file_b

def render_tolerance_section():
    with st.container(border=True):
        st.markdown(
            """
            <div class="panel-title" style="min-height:64px;padding:0;border-bottom:1px solid #eeedf3;margin-bottom:18px;">
                <div>
                    <h3>Tolerância financeira</h3>
                    <div class="panel-subtitle">Defina o limite aceito para diferenças de arredondamento.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        opts = {"R$ 0,00": 0.0, "R$ 0,30": 0.30, "R$ 0,50": 0.50, "R$ 1,00": 1.0, "Personalizado": -1}
        selected = st.radio("Tolerância", list(opts.keys()), index=2, horizontal=True, label_visibility="collapsed")
        value = opts[selected]
        if value == -1:
            value = st.number_input("Valor personalizado (R$)", 0.0, 999.0, 0.50, 0.10, format="%.2f")
    return value


def render_ops_panel():
    hist = auditoria_io.carregar_historico()
    latest = hist[0] if hist else None
    current = None
    if st.session_state.get("df_res") is not None and not st.session_state["df_res"].empty:
        resumo = st.session_state.get("resumo", {})
        current = {
            "arquivo_a": st.session_state.get("nome_a", "-"),
            "arquivo_b": st.session_state.get("nome_b", "-"),
            "tolerancia": st.session_state.get("tol", 0),
            "total": resumo.get("total", 0),
            "divergentes": resumo.get("divergentes", 0),
            "faltantes": resumo.get("faltantes_a", 0) + resumo.get("faltantes_b", 0),
            "impacto": resumo.get("impacto_absoluto", 0),
        }

    html_parts = [
        '<div class="panel span-12">',
        '<div class="panel-title"><div><h3>Visão operacional</h3><div class="panel-subtitle">Resumo da sessão atual e da última auditoria salva.</div></div></div>',
        '<div class="panel-body">',
    ]

    if current:
        html_parts.append(
            f'<div class="history-grid ops-grid">'
            f'<div class="history-metric"><small>Sessão atual</small><b>{safe_text(current["arquivo_a"])} x {safe_text(current["arquivo_b"])}</b></div>'
            f'<div class="history-metric"><small>Total analisado</small><b>{int(current["total"] or 0)}</b></div>'
            f'<div class="history-metric"><small>Divergentes</small><b>{int(current["divergentes"] or 0)}</b></div>'
            f'<div class="history-metric"><small>Faltantes</small><b>{int(current["faltantes"] or 0)}</b></div>'
            f'</div>'
        )
        html_parts.append(
            f'<div class="ops-note"><b>Tolerância atual:</b> {br_money(current["tolerancia"])}<br><b>Impacto crítico atual:</b> {br_money(current["impacto"])}<br><b>Arquivos em análise:</b> {safe_text(current["arquivo_a"])} e {safe_text(current["arquivo_b"])}.</div>'
        )
    elif latest:
        html_parts.append(
            f'<div class="history-grid ops-grid">'
            f'<div class="history-metric"><small>Última auditoria</small><b>{safe_text(latest.get("data_hora", "-"))}</b></div>'
            f'<div class="history-metric"><small>Total analisado</small><b>{int(latest.get("total", 0) or 0)}</b></div>'
            f'<div class="history-metric"><small>Divergentes</small><b>{int(latest.get("divergentes", 0) or 0)}</b></div>'
            f'<div class="history-metric"><small>Faltantes</small><b>{int(latest.get("faltantes_a", 0) or 0) + int(latest.get("faltantes_b", 0) or 0)}</b></div>'
            f'</div>'
        )
        html_parts.append(
            f'<div class="ops-note"><b>Total de auditorias salvas:</b> {len(hist)}<br><b>Última tolerância:</b> {br_money(latest.get("tolerancia", 0))}<br><b>Último impacto crítico:</b> {br_money(latest.get("impacto_absoluto", 0))}</div>'
        )
    else:
        html_parts.append('<div class="ops-note"><b>Nenhuma auditoria processada nesta sessão.</b><br>Envie dois relatórios válidos para gerar o primeiro comparativo.</div>')

    html_parts.append('</div></div>')
    st.markdown(''.join(html_parts), unsafe_allow_html=True)
def kpi_html(label, value, tone="purple", money=False):
    money_class = " money" if money else ""
    return (
        f'<div class="panel span-3">'
        f'<div class="panel-body">'
        f'<div class="kpi-label">{safe_text(label)}</div>'
        f'<div class="kpi-value{money_class} tone-{tone}">{safe_text(value)}</div>'
        f'</div>'
        f'</div>'
    )


def render_kpis(resumo, df):
    df_calc = df.copy()
    df_calc["EmpA"] = df_calc["Empresa A"].fillna(0)
    df_calc["EmpB"] = df_calc["Empresa B"].fillna(0)
    df_calc["MotA"] = df_calc["Motorista A"].fillna(0)
    df_calc["MotB"] = df_calc["Motorista B"].fillna(0)
    df_calc["DE"] = df_calc["EmpA"] - df_calc["EmpB"]
    df_calc["DM"] = df_calc["MotA"] - df_calc["MotB"]
    df_calc["MD"] = df_calc[["DE", "DM"]].abs().max(axis=1)

    div_df = df_calc[df_calc["Status"] == "Divergente"]
    crit_df = df_calc[df_calc["Status"].isin(["Divergente", "Faltante no A", "Faltante no B"])]

    de_div = div_df["DE"].sum()
    dm_div = div_df["DM"].sum()
    imp_div = div_df["MD"].sum()
    de_crit = crit_df["DE"].sum()
    dm_crit = crit_df["DM"].sum()
    imp_geral = df_calc["MD"].sum()

    summary_html = (
        '<div class="breadcrumb"><b>Resumo da auditoria</b><span>Resultado consolidado dos documentos processados</span></div>'
        '<div class="dashboard-grid">'
        f'{kpi_html("Total analisado", resumo["total"], "purple")}'
        f'{kpi_html("OK", resumo["ok"], "green")}'
        f'{kpi_html("OK Arred.", resumo["ok_arredondamento"], "blue")}'
        f'{kpi_html("Diferenças dentro da tolerância", resumo["ok_arredondamento"], "blue")}'
        f'{kpi_html("Divergentes reais", resumo["divergentes"], "red")}'
        f'{kpi_html("Faltantes no A", resumo["faltantes_a"], "orange")}'
        f'{kpi_html("Faltantes no B", resumo["faltantes_b"], "orange")}'
        f'{kpi_html("Impacto Crítico Total", br_money(resumo["impacto_absoluto"]), "red", True)}'
        '</div>'
        '<div class="breadcrumb" style="margin-top:24px;"><b>Leitura financeira complementar</b><span>Detalhes de apoio para análise executiva</span></div>'
        '<div class="dashboard-grid">'
        f'{kpi_html("Diferença Empresa - Divergentes", br_money(de_div), "blue", True)}'
        f'{kpi_html("Diferença Motorista - Divergentes", br_money(dm_div), "purple", True)}'
        f'{kpi_html("Impacto - Divergentes", br_money(imp_div), "red", True)}'
        f'{kpi_html("Diferença Empresa - Crítico Total", br_money(de_crit), "blue", True)}'
        f'{kpi_html("Diferença Motorista - Crítico Total", br_money(dm_crit), "purple", True)}'
        '</div>'
        '<div class="panel span-12" style="margin-top:24px;">'
        '<div class="panel-body">'
        f'<div class="impact-line">Impacto geral incluindo arredondamentos: <b>{br_money(imp_geral)}</b></div>'
        '<div class="explain">'
        '<b>Divergentes</b> considera apenas CTEs encontrados nos dois relatórios com diferença acima da tolerância. '
        '<b>Crítico total</b> soma divergentes reais e CTEs faltantes. '
        '<b>OK Arred.</b> representa diferenças pequenas dentro da tolerância configurada. '
        'As diferenças dentro da tolerância continuam visíveis para conferência, mas não entram no impacto crítico.'
        '</div>'
        '</div>'
        '</div>'
    )
    st.markdown(summary_html, unsafe_allow_html=True)
def render_chart(df):
    with st.container(border=True):
        st.markdown(
            """
            <div class="panel-title" style="min-height:64px;padding:0;border-bottom:1px solid #eeedf3;margin-bottom:18px;">
                <div>
                    <h3>Distribuição por status</h3>
                    <div class="panel-subtitle">Volume de CTEs por classificação.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        counts = df["Status"].value_counts().reset_index()
        counts.columns = ["Status", "Qtd"]
        counts["Status visual"] = counts["Status"].replace({"OK por arredondamento": "OK Arred."})
        colors = {
            "OK": "#22c55e",
            "OK Arred.": "#179cf4",
            "Divergente": "#ef4444",
            "Faltante no A": "#ff9f1c",
            "Faltante no B": "#ff9f1c",
        }
        fig = px.bar(
            counts,
            x="Qtd",
            y="Status visual",
            color="Status visual",
            orientation="h",
            text="Qtd",
            color_discrete_map=colors,
        )
        fig.update_layout(
            height=270,
            margin=dict(l=6, r=12, t=6, b=6),
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            showlegend=False,
            font=dict(family="Manrope, Segoe UI, sans-serif", color="#202236"),
            xaxis=dict(title="", gridcolor="#efedf6", zeroline=False),
            yaxis=dict(title="", showgrid=False),
        )
        fig.update_traces(marker_line_width=0, textposition="outside", hovertemplate="%{y}: %{x}<extra></extra>")
        st.plotly_chart(fig, use_container_width=True)
def status_style(value):
    styles = {
        "OK": "background-color:#dcfce7;color:#166534;font-weight:800;border-radius:99px;",
        "OK Arred.": "background-color:#e8f3ff;color:#075985;font-weight:800;border-radius:99px;",
        "Divergente": "background-color:#fee2e2;color:#991b1b;font-weight:800;border-radius:99px;",
        "Faltante no A": "background-color:#fef3c7;color:#92400e;font-weight:800;border-radius:99px;",
        "Faltante no B": "background-color:#fef3c7;color:#92400e;font-weight:800;border-radius:99px;",
    }
    return styles.get(value, "")


def status_badge_html(value):
    tone = {
        "OK": "ok",
        "OK Arred.": "round",
        "Divergente": "div",
        "Faltante no A": "miss",
        "Faltante no B": "miss",
    }.get(value, "round")
    return f'<span class="table-badge {tone}">{safe_text(value)}</span>'


def get_table_row_class(row_meta):
    if row_meta.get("_Linha Visual") == "missing":
        return "row-missing"
    return ""


def get_table_cell_class(column, row_meta, money_columns):
    classes = []
    if column in money_columns:
        classes.append("cell-money")
    elif column == "Observação":
        classes.append("cell-observation")
    elif column == "CTE":
        classes.append("cell-cte")

    if row_meta.get("_Linha Visual") != "missing":
        if column in {"Empresa A", "Empresa B", "Dif. Empresa"}:
            if row_meta.get("_Empresa Visual") == "critical":
                classes.append("cell-diff-critical")
            elif row_meta.get("_Empresa Visual") == "tolerance":
                classes.append("cell-diff-tolerance")
        elif column in {"Motorista A", "Motorista B", "Diferença"}:
            if row_meta.get("_Motorista Visual") == "critical":
                classes.append("cell-diff-critical")
            elif row_meta.get("_Motorista Visual") == "tolerance":
                classes.append("cell-diff-tolerance")

    return " ".join(classes)


def render_table(df):
    tolerancia_visual = resolve_tolerance_value(st.session_state.get("tol", 0.50))
    prepared = prepare_conference_dataframe(df, tolerancia_visual)

    with st.container(border=True):
        st.markdown(
            """
            <div class="section-head" style="margin-bottom:14px;">
                <div>
                    <h3 class="section-title">Conferência detalhada das diferenças</h3>
                    <div class="section-subtitle">Visualize divergentes, arredondamentos e faltantes com destaque apenas no campo que divergiu.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        resumo = {
            "divergentes": int((prepared["_Status Base"] == "Divergente").sum()),
            "ok_arredondamento": int((prepared["_Status Base"] == "OK por arredondamento").sum()),
            "faltantes_a": int((prepared["_Status Base"] == "Faltante no A").sum()),
            "faltantes_b": int((prepared["_Status Base"] == "Faltante no B").sum()),
        }
        st.markdown(build_detailed_counts_html(resumo), unsafe_allow_html=True)

        c1, c2, c3, c4 = st.columns([1.1, 1, 1, 1])
        with c1:
            search = st.text_input("Buscar CTE", placeholder="Ex.: 1752")
        with c2:
            scope = st.selectbox("Status", CONFERENCE_FILTER_LABELS)
        with c3:
            diff_label = st.selectbox("Diferença mínima", list(VISUAL_DIFF_FILTERS.keys()), index=0)
            if diff_label == "Personalizado":
                diff = st.number_input("Valor personalizado (R$)", min_value=0.0, value=0.01, step=0.01, format="%.2f")
            else:
                diff = VISUAL_DIFF_FILTERS[diff_label]
        with c4:
            order = st.selectbox("Ordenar por", ["Maior diferença", "CTE crescente", "Status"])

        visible_prepared = apply_conference_filters(prepared, scope, diff, search, order)
        visible, money_columns = build_conference_display_table(visible_prepared)

        header_html = ''.join(
            f'<th class="{"cell-money" if column in money_columns else "cell-observation" if column == "Observação" else "cell-cte" if column == "CTE" else ""}">{safe_text(column)}</th>'
            for column in visible.columns
        )

        row_html = []
        for row_index, row in visible.iterrows():
            row_meta = visible_prepared.loc[row_index]
            cells = []
            for column in visible.columns:
                value = row[column]
                if column == "Status":
                    cells.append(f'<td>{status_badge_html(value)}</td>')
                else:
                    css_class = get_table_cell_class(column, row_meta, money_columns)
                    cells.append(f'<td class="{css_class}">{safe_text(value)}</td>')
            row_html.append(f'<tr class="{get_table_row_class(row_meta)}">{"".join(cells)}</tr>')

        if not row_html:
            row_html.append(f'<tr><td class="cell-empty" colspan="{len(visible.columns)}">Nenhum registro encontrado com os filtros atuais.</td></tr>')

        table_html = '<div class="table-shell audit-results-shell"><table class="audit-table">' + f'<thead><tr>{header_html}</tr></thead>' + f'<tbody>{"".join(row_html)}</tbody></table></div>' + '<div class="table-scroll-hint">Use a roda do mouse sobre a tabela para rolar sem descer a página inteira.</div>'
        st.markdown(table_html, unsafe_allow_html=True)
def render_exports(df, resumo, name_a, name_b, tolerance):
    render_breadcrumb("Exportar relatório", "Arquivos executivos")
    c1, c2, c3, c4 = st.columns(4, gap="large")
    with c1:
        with st.container(border=True):
            st.markdown('<div class="export-card"><h4>CSV</h4><p>Dados brutos para análise.</p></div>', unsafe_allow_html=True)
            st.download_button("Baixar CSV", auditoria_io.exportar_csv(df), "auditoria.csv", "text/csv", use_container_width=True)
    with c2:
        with st.container(border=True):
            st.markdown('<div class="export-card"><h4>Excel</h4><p>Planilha formatada.</p></div>', unsafe_allow_html=True)
            st.download_button("Baixar Excel", build_excel_bytes(df, resumo, name_a, name_b, tolerance), "auditoria.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with c3:
        with st.container(border=True):
            st.markdown('<div class="export-card"><h4>PDF Executivo</h4><p>Resumo executivo com foco em criticidade.</p></div>', unsafe_allow_html=True)
            st.download_button("Baixar PDF Executivo", build_executive_pdf_bytes(df, resumo, name_a, name_b, tolerance), "auditoria_executiva.pdf", "application/pdf", use_container_width=True)
    with c4:
        with st.container(border=True):
            st.markdown('<div class="export-card"><h4>PDF de Conferência Detalhada</h4><p>Resumo, critérios, tolerância, críticos, diferenças dentro da tolerância, faltantes e tabela completa.</p></div>', unsafe_allow_html=True)
            st.download_button("Baixar PDF de Conferência Detalhada", build_detailed_pdf_bytes(df, resumo, name_a, name_b, tolerance), "auditoria_conferencia_detalhada.pdf", "application/pdf", use_container_width=True)
def render_history_page():
    render_topbar("Histórico")
    render_breadcrumb("FreteScan", "Histórico de auditorias")
    hist = auditoria_io.carregar_historico()
    if not hist:
        st.markdown(
            """
            <div class="panel span-12">
                <div class="panel-title"><div><h3>Histórico</h3><div class="panel-subtitle">Nenhuma auditoria registrada.</div></div></div>
                <div class="panel-body"><div class="explain">Nenhuma auditoria registrada.</div></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    for item in pd.DataFrame(hist).head(8).to_dict("records"):
        st.markdown(
            f"""
            <div class="panel span-12">
                <div class="panel-title">
                    <div>
                        <h3>{safe_text(item.get('arquivo_a', ''))} x {safe_text(item.get('arquivo_b', ''))}</h3>
                        <div class="panel-subtitle">{safe_text(item.get("data_hora", ""))}</div>
                    </div>
                </div>
                <div class="panel-body">
                    <div class="history-grid">
                        <div class="history-metric"><small>Tolerância</small><b>{br_money(item.get("tolerancia", 0))}</b></div>
                        <div class="history-metric"><small>Total</small><b>{int(item.get("total", 0))}</b></div>
                        <div class="history-metric"><small>Divergentes</small><b>{int(item.get("divergentes", 0))}</b></div>
                        <div class="history-metric"><small>Faltantes</small><b>{int(item.get("faltantes_a", 0)) + int(item.get("faltantes_b", 0))}</b></div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_reports_page():
    render_topbar("Relatórios")
    render_breadcrumb("FreteScan", "Relatórios")
    hist = auditoria_io.carregar_historico()

    if not hist:
        st.markdown(
            """
            <div class="panel span-12">
                <div class="panel-title"><div><h3>Relatórios</h3><div class="panel-subtitle">Nenhuma auditoria disponível para análise.</div></div></div>
                <div class="panel-body"><div class="explain">Execute uma auditoria para começar a preencher esta área com histórico consolidado.</div></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    hist_df = pd.DataFrame(hist).fillna(0)
    hist_df["faltantes_total"] = hist_df.get("faltantes_a", 0) + hist_df.get("faltantes_b", 0)
    hist_df["arquivos"] = hist_df["arquivo_a"].astype(str) + " x " + hist_df["arquivo_b"].astype(str)

    total_auditorias = len(hist_df)
    media_total = int(round(hist_df.get("total", pd.Series(dtype=float)).mean())) if "total" in hist_df else 0
    media_div = int(round(hist_df.get("divergentes", pd.Series(dtype=float)).mean())) if "divergentes" in hist_df else 0
    ultima_execucao = str(hist_df.iloc[0].get("data_hora", "-"))

    st.markdown(
        f"""
        <div class="dashboard-grid">
            {kpi_html("Auditorias salvas", total_auditorias, "purple")}
            {kpi_html("Média total analisado", media_total, "blue")}
            {kpi_html("Média divergentes", media_div, "red")}
            {kpi_html("Última execução", ultima_execucao, "green")}
        </div>
        """,
        unsafe_allow_html=True,
    )

    search_col, tol_col = st.columns([1.4, 1])
    with search_col:
        search = st.text_input("Buscar por arquivo", placeholder="Ex.: ABRIL ATUA")
    with tol_col:
        tolerancias = ["Todas"] + [f"R$ {float(x):.2f}" for x in sorted(hist_df["tolerancia"].astype(float).unique())]
        selected_tol = st.selectbox("Tolerância", tolerancias)

    visible = hist_df.copy()
    if search:
        search_norm = search.strip().lower()
        visible = visible[
            visible["arquivo_a"].astype(str).str.lower().str.contains(search_norm, na=False)
            | visible["arquivo_b"].astype(str).str.lower().str.contains(search_norm, na=False)
        ]
    if selected_tol != "Todas":
        visible = visible[visible["tolerancia"].astype(float) == float(selected_tol.replace("R$", "").strip())]

    visible = visible[[
        "data_hora",
        "arquivo_a",
        "arquivo_b",
        "tolerancia",
        "total",
        "divergentes",
        "faltantes_total",
    ]].rename(columns={
        "data_hora": "Data/Hora",
        "arquivo_a": "Arquivo A",
        "arquivo_b": "Arquivo B",
        "tolerancia": "Tolerância",
        "total": "Total analisado",
        "divergentes": "Divergentes",
        "faltantes_total": "Faltantes",
    })

    visible["Tolerância"] = visible["Tolerância"].apply(br_money)

    header_html = ''.join(f'<th>{safe_text(col)}</th>' for col in visible.columns)
    rows = []
    for _, row in visible.iterrows():
        cells = []
        for col in visible.columns:
            css_class = 'cell-money' if col in ["Tolerância"] else ''
            cells.append(f'<td class="{css_class}">{safe_text(row[col])}</td>')
        rows.append(f'<tr>{"".join(cells)}</tr>')
    if not rows:
        rows.append(f'<tr><td class="cell-empty" colspan="{len(visible.columns)}">Nenhum relatório encontrado com os filtros atuais.</td></tr>')

    st.markdown(
        """
        <div class="panel-title" style="min-height:64px;padding:0;border-bottom:1px solid #eeedf3;margin:28px 0 18px 0;">
            <div>
                <h3>Relatórios salvos</h3>
                <div class="panel-subtitle">Consulte auditorias já executadas com filtro por arquivo e tolerância.</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f'<div class="table-shell"><table class="audit-table"><thead><tr>{header_html}</tr></thead><tbody>{"".join(rows)}</tbody></table></div>',
        unsafe_allow_html=True,
    )

    if st.session_state.get("df_res") is not None and not st.session_state["df_res"].empty:
        st.markdown(
            """
            <div class="panel-title" style="min-height:64px;padding:0;border-bottom:1px solid #eeedf3;margin:28px 0 18px 0;">
                <div>
                    <h3>Relatório atual</h3>
                    <div class="panel-subtitle">Exportações da auditoria aberta nesta sessão.</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        render_exports(
            st.session_state["df_res"],
            st.session_state["resumo"],
            st.session_state["nome_a"],
            st.session_state["nome_b"],
            st.session_state["tol"],
        )

def render_about_page():
    render_topbar("Sobre")
    render_breadcrumb("FreteScan", "Sobre")
    st.markdown(
        f"""
        <div class="panel span-12">
            <div class="panel-title"><div><h3>{safe_text(BRAND_NAME)}</h3><div class="panel-subtitle">{safe_text(BRAND_PLATFORM)}</div></div></div>
            <div class="panel-body">
                <p class="explain">
                    A {safe_text(BRAND_NAME)} foi desenhada para transformar auditoria de frete em uma frente de inteligência logística.
                    A plataforma cruza relatórios por CTE, evidencia impacto financeiro, separa arredondamentos e organiza a operação com uma leitura mais executiva,
                    próxima da comunicação comercial do seu site de referência e coerente com a identidade da marca.
                </p>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_placeholder(page):
    render_topbar(page)
    render_breadcrumb("FreteScan", page)
    st.markdown(
        f"""
        <div class="panel span-12">
            <div class="panel-title"><div><h3>{safe_text(page)}</h3><div class="panel-subtitle">Módulo em preparação.</div></div></div>
            <div class="panel-body"><div class="explain">Esta funcionalidade será disponibilizada em breve.</div></div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_settings_page():
    render_topbar("Configurações")
    render_breadcrumb("FreteScan", "Configurações")
    st.markdown(
        """
        <div class="dashboard-grid">
            <div class="panel span-6">
                <div class="panel-title"><div><h3>Ambiente</h3><div class="panel-subtitle">Informações operacionais da instância atual.</div></div></div>
                <div class="panel-body">
                    <div class="history-grid">
                        <div class="history-metric"><small>Status</small><b>Online</b></div>
                        <div class="history-metric"><small>Versão</small><b>v1.0.0</b></div>
                        <div class="history-metric"><small>Interface</small><b>Streamlit</b></div>
                        <div class="history-metric"><small>Persistência</small><b>Histórico local</b></div>
                    </div>
                </div>
            </div>
            <div class="panel span-6">
                <div class="panel-title"><div><h3>Regras visíveis</h3><div class="panel-subtitle">Parâmetros operacionais expostos ao usuário.</div></div></div>
                <div class="panel-body">
                    <div class="history-grid">
                        <div class="history-metric"><small>Tolerâncias</small><b>0,00 | 0,30 | 0,50 | 1,00</b></div>
                        <div class="history-metric"><small>Formatos</small><b>PDF | Excel | CSV</b></div>
                        <div class="history-metric"><small>Comparação</small><b>Por CTE</b></div>
                        <div class="history-metric"><small>Histórico</small><b>Até 100 auditorias</b></div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_processing_card(current_step=1, current_label=None):
    current_step = max(1, min(current_step, len(PROCESSING_STEPS)))
    active_title = current_label or PROCESSING_STEPS[current_step - 1][0]
    steps_markup = []
    for index, (title, copy) in enumerate(PROCESSING_STEPS, start=1):
        if index < current_step:
            state_class = "step-done"
            state_label = "Concluída"
        elif index == current_step:
            state_class = "step-active"
            state_label = "Etapa atual"
        else:
            state_class = "step-pending"
            state_label = "Pendente"
        steps_markup.append(
            (
                f'<div class="processing-step {state_class}">'
                '<div class="processing-step-top">'
                f'<span class="processing-step-state">{safe_text(state_label)}</span>'
                f'<span class="processing-step-index">{index}</span>'
                '</div>'
                '<div>'
                f'<div class="processing-step-title">{safe_text(title)}</div>'
                f'<div class="processing-step-copy">{safe_text(copy)}</div>'
                '</div>'
                '</div>'
            )
        )

    card_html = (
        '<div class="processing-shell">'
        '<div class="processing-head">'
        '<div class="processing-kicker">Processamento em andamento</div>'
        '<div class="processing-title">Processando auditoria</div>'
        '<div class="processing-subtitle">Leitura dos dois relatórios, cruzamento dos CTEs, aplicação da tolerância e montagem do relatório final.</div>'
        '<div class="processing-stage-note">'
        '<span class="processing-stage-dot"></span>'
        f'<span>Etapa atual: {safe_text(active_title)}</span>'
        '</div>'
        '</div>'
        f'<div class="processing-grid">{"".join(steps_markup)}</div>'
        '</div>'
    )
    st.markdown(card_html, unsafe_allow_html=True)


def update_processing_view(status_box, current_step, progress_value, progress_text):
    status_box.empty()
    with status_box.container():
        render_processing_card(current_step=current_step, current_label=progress_text)
        st.progress(progress_value, text=progress_text)

def clear_audit_workspace():
    clear_export_caches()
    reset_audit_state()
    for key in list(st.session_state.keys()):
        if key.startswith("up_a_widget_") or key.startswith("up_b_widget_"):
            st.session_state.pop(key, None)
    for prefix in ["up_a", "up_b"]:
        st.session_state.pop(f"{prefix}_stored", None)
        st.session_state.pop(prefix, None)
        version_key = f"{prefix}_widget_version"
        st.session_state[version_key] = int(st.session_state.get(version_key, 0) or 0) + 1
    st.session_state.pop("audit_error", None)
    st.session_state.pop("audit_error_details", None)
    st.session_state.pop("audit_debug", None)
    st.session_state.pop("caminho_a_temp", None)
    st.session_state.pop("caminho_b_temp", None)
    st.session_state.processing = False


def render_app_header():
    st.markdown(
        f"""
        <div class="compact-header-shell">
            <div class="compact-header">
                <div style="display:flex;align-items:center;justify-content:space-between;gap:16px;flex-wrap:wrap;">
                    <div class="compact-brand">
                        <div class="compact-brand-mark">{fretevision_mark(48)}</div>
                        <div>
                            <div class="compact-brand-title">FRETE<strong>VISION</strong></div>
                            <div class="compact-brand-subtitle">{safe_text(BRAND_TAGLINE)}</div>
                        </div>
                    </div>
                    <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;">
                        <span class="upload-ready-chip">Nova Auditoria</span>
                        <span class="header-status-pill">Online</span>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_compact_menu():
    pages = ["Auditoria", "Visão que move resultados", "Suporte Técnico", "Novidades"]
    if "main_page" not in st.session_state or st.session_state.main_page not in pages:
        st.session_state.main_page = "Auditoria"
    st.markdown('<div class="menu-shell"><div class="menu-caption">Menu</div></div>', unsafe_allow_html=True)
    return st.radio(
        "Menu principal",
        pages,
        key="main_page",
        horizontal=True,
        label_visibility="collapsed",
    )


def render_tolerance_section():
    with st.container(border=True):
        head_left, head_right = st.columns([7, 1.35], gap="small")
        with head_left:
            st.markdown(
                """
                <div class="section-head">
                    <h3 class="section-title">Nova Auditoria</h3>
                    <div class="section-subtitle">Configure a tolerância e envie os dois relatórios do mesmo período.</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with head_right:
            if st.button("Limpar", key="clear_workspace", use_container_width=True):
                clear_audit_workspace()
                st.rerun()

        label_col, radio_col = st.columns([1.6, 6], gap="small")
        with label_col:
            st.markdown('<div class="compact-field-label">Tolerância de diferença</div>', unsafe_allow_html=True)
        with radio_col:
            opts = {"R$ 0,00": 0.0, "R$ 0,30": 0.30, "R$ 0,50": 0.50, "R$ 1,00": 1.0, "Personalizado": -1}
            selected = st.radio("Tolerância", list(opts.keys()), index=2, horizontal=True, label_visibility="collapsed")
            value = opts[selected]
            if value == -1:
                value = st.number_input("Valor personalizado (R$)", 0.0, 999.0, 0.50, 0.10, format="%.2f")
        return value


def render_upload_box(title, description, key):
    stored_key = f"{key}_stored"
    version_key = f"{key}_widget_version"
    legacy_upload = st.session_state.get(key)
    if st.session_state.get(stored_key) is None and legacy_upload is not None and hasattr(legacy_upload, "getvalue"):
        st.session_state[stored_key] = serialize_uploaded_file(legacy_upload)
        st.session_state.pop(key, None)

    current_file = st.session_state.get(stored_key)
    widget_version = int(st.session_state.get(version_key, 0) or 0)
    widget_key = f"{key}_widget_{widget_version}"

    with st.container(border=True):
        st.markdown(
            f"""
            <div class="upload-panel-head">
                <div class="upload-panel-icon">{icon_doc(18)}</div>
                <div>
                    <div class="upload-panel-title">{safe_text(title)}</div>
                    <div class="upload-panel-subtitle">{safe_text(description)}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if current_file is None:
            uploaded = st.file_uploader(title, type=["pdf"], key=widget_key, label_visibility="collapsed")
            if uploaded:
                st.session_state[stored_key] = serialize_uploaded_file(uploaded)
                st.session_state.pop(widget_key, None)
                st.rerun()
            return None

        st.markdown(
            f"""
            <div class="upload-ready-bar">
                <span class="upload-ready-state"><span class="upload-ready-dot"></span>Arquivo carregado</span>
                <span class="upload-ready-chip">{safe_text(get_upload_name(current_file))} • {safe_text(file_size(current_file))}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        action_col_a, action_col_b = st.columns(2, gap="small")
        with action_col_a:
            if st.button("Trocar arquivo", key=f"{key}_replace", use_container_width=True):
                st.session_state.pop(stored_key, None)
                st.session_state.pop(widget_key, None)
                st.session_state[version_key] = widget_version + 1
                reset_audit_state()
                st.session_state.audit_error = None
                st.session_state.audit_error_details = []
                st.rerun()
        with action_col_b:
            if st.button("Remover arquivo", key=f"{key}_remove", use_container_width=True):
                st.session_state.pop(stored_key, None)
                st.session_state.pop(widget_key, None)
                st.session_state[version_key] = widget_version + 1
                reset_audit_state()
                st.session_state.audit_error = None
                st.session_state.audit_error_details = []
                st.rerun()
        return current_file


def render_upload_section():
    col_a, col_b = st.columns(2, gap="medium")
    with col_a:
        file_a = render_upload_box("Sistema A (Relatório DL)", "Upload do relatório principal da empresa.", "up_a")
    with col_b:
        file_b = render_upload_box("Sistema B (Relatório Carreteiro)", "Upload do relatório de conferência.", "up_b")
    return file_a, file_b


def render_kpis(resumo, df):
    faltantes_total = int(resumo["faltantes_a"]) + int(resumo["faltantes_b"])
    tolerance_text = br_money(st.session_state.get("tol", 0.50))
    summary_html = (
        '<div class="section-head" style="margin-top:10px;">'
        '<h3 class="section-title">Resumo da auditoria</h3>'
        '<div class="section-subtitle">Resultado direto da comparação entre os dois relatórios processados.</div>'
        '</div>'
        '<div class="summary-grid">'
        f'<div class="summary-card"><small>Total analisado</small><b>{safe_text(resumo["total"])}</b><span>Tolerância: {safe_text(tolerance_text)}</span></div>'
        f'<div class="summary-card ok"><small>OK</small><b>{safe_text(resumo["ok"])}</b><span>Sem diferença identificada</span></div>'
        f'<div class="summary-card round"><small>OK Arred.</small><b>{safe_text(resumo["ok_arredondamento"])}</b><span>Diferenças dentro da tolerância</span></div>'
        f'<div class="summary-card div"><small>Divergentes reais</small><b>{safe_text(resumo["divergentes"])}</b><span>Acima da tolerância configurada</span></div>'
        f'<div class="summary-card miss"><small>Faltantes</small><b>{safe_text(faltantes_total)}</b><span>No A: {safe_text(resumo["faltantes_a"])} • No B: {safe_text(resumo["faltantes_b"])}</span></div>'
        f'<div class="summary-card impact"><small>Impacto crítico total</small><b>{safe_text(br_money(resumo["impacto_absoluto"]))}</b><span>Não inclui OK por arredondamento</span></div>'
        '</div>'
        '<div class="summary-note">Diferenças dentro da tolerância continuam visíveis para conferência na tabela, mas não compõem o impacto crítico.</div>'
    )
    st.markdown(summary_html, unsafe_allow_html=True)


def render_chart(df):
    counts = (
        df["Status"]
        .astype(str)
        .replace({"OK por arredondamento": "OK Arred."})
        .value_counts()
        .reindex(["OK", "OK Arred.", "Divergente", "Faltante no A", "Faltante no B"], fill_value=0)
        .rename_axis("Status")
        .reset_index(name="Qtd")
    )
    counts = counts[counts["Qtd"] > 0].copy()
    if counts.empty:
        return

    colors = {
        "OK": "#16A34A",
        "OK Arred.": "#2563EB",
        "Divergente": "#DC2626",
        "Faltante no A": "#F97316",
        "Faltante no B": "#FB923C",
    }
    total = int(counts["Qtd"].sum())
    fig = px.pie(
        counts,
        names="Status",
        values="Qtd",
        color="Status",
        hole=0.72,
        color_discrete_map=colors,
    )
    fig.update_traces(
        sort=False,
        direction="clockwise",
        textinfo="none",
        hovertemplate="<b>%{label}</b><br>%{value} CTEs<extra></extra>",
        marker=dict(line=dict(color="#ffffff", width=3)),
    )
    fig.update_layout(
        template="plotly_white",
        height=300,
        margin=dict(l=10, r=10, t=8, b=44),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#111827", size=13),
        legend=dict(
            orientation="h",
            y=-0.08,
            x=0.0,
            yanchor="top",
            xanchor="left",
            title_text="",
            font=dict(color="#111827", size=12),
        ),
        annotations=[
            dict(
                text=f"<b>{total}</b><br>Total",
                x=0.5,
                y=0.5,
                showarrow=False,
                font=dict(color="#111827", size=16),
            )
        ],
    )

    with st.container(border=True):
        st.markdown(
            """
            <div class="section-head chart-shell">
                <h3 class="section-title">Status da auditoria</h3>
                <div class="section-subtitle">Leitura rápida da distribuição por classificação, com foco em conferência e visibilidade.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def render_exports(df, resumo, name_a, name_b, tolerance):
    st.markdown(
        """
        <div class="section-head export-actions">
            <h3 class="section-title">Exportações</h3>
            <div class="export-note">Baixe CSV, Excel e os dois PDFs no mesmo padrão visual da conferência.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    c1, c2, c3, c4 = st.columns(4, gap="small")
    with c1:
        st.download_button("Baixar CSV", auditoria_io.exportar_csv(df), "auditoria.csv", "text/csv", use_container_width=True)
    with c2:
        st.download_button("Baixar Excel", build_excel_bytes(df, resumo, name_a, name_b, tolerance), "auditoria.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with c3:
        st.download_button("Baixar PDF Executivo", build_executive_pdf_bytes(df, resumo, name_a, name_b, tolerance), "auditoria_executiva.pdf", "application/pdf", use_container_width=True)
    with c4:
        st.download_button("Baixar PDF de Conferência Detalhada", build_detailed_pdf_bytes(df, resumo, name_a, name_b, tolerance), "auditoria_conferencia_detalhada.pdf", "application/pdf", use_container_width=True)


def render_marketing_page():
    st.markdown(
        f"""
        <div class="marketing-hero">
            <div class="marketing-card">
                <div class="marketing-kicker">Frete Vision</div>
                <div class="marketing-title">Visão que move resultados.</div>
                <div class="marketing-copy">
                    Uma plataforma criada para transformar conferência logística em leitura executiva clara.
                    O produto evidencia divergências por CTE, revela impacto real e entrega auditoria pronta para operação,
                    apresentação comercial e amostragem com a cliente.
                </div>
                <div class="marketing-chip-row">
                    <span class="marketing-chip">Comparação precisa por CTE</span>
                    <span class="marketing-chip">Conferência visual por campo</span>
                    <span class="marketing-chip">PDF executivo e detalhado</span>
                    <span class="marketing-chip">Diferenças de centavos visíveis</span>
                </div>
                <div class="summary-note" style="margin-top:18px;">
                    A aba <b>Auditoria</b> continua isolada para uso operacional. Esta área fica dedicada à apresentação do produto.
                </div>
            </div>
            <div class="marketing-visual">
                <div class="visual-glow"></div>
                <div class="visual-signature">{fretevision_signature()}</div>
                <div class="visual-float-row">
                    <div class="visual-stat">
                        <small>Leitura</small>
                        <b>Clara e imediata</b>
                    </div>
                    <div class="visual-stat">
                        <small>Conferência</small>
                        <b>CTE por CTE</b>
                    </div>
                    <div class="visual-stat">
                        <small>Saídas</small>
                        <b>PDF + Excel</b>
                    </div>
                </div>
                <div class="visual-board">
                    <div class="visual-board-title">Visão resumida da operação</div>
                    <div class="visual-bars">
                        <div class="visual-bar-row">
                            <div class="visual-bar-label">OK</div>
                            <div class="visual-bar-track"><div class="visual-bar-fill" style="width:72%;background:#22c55e;"></div></div>
                            <div class="visual-bar-value">72</div>
                        </div>
                        <div class="visual-bar-row">
                            <div class="visual-bar-label">OK Arred.</div>
                            <div class="visual-bar-track"><div class="visual-bar-fill" style="width:54%;background:#3b82f6;"></div></div>
                            <div class="visual-bar-value">54</div>
                        </div>
                        <div class="visual-bar-row">
                            <div class="visual-bar-label">Diverg.</div>
                            <div class="visual-bar-track"><div class="visual-bar-fill" style="width:41%;background:#ef4444;"></div></div>
                            <div class="visual-bar-value">41</div>
                        </div>
                        <div class="visual-bar-row">
                            <div class="visual-bar-label">Faltantes</div>
                            <div class="visual-bar-track"><div class="visual-bar-fill" style="width:18%;background:#f59e0b;"></div></div>
                            <div class="visual-bar-value">18</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        <div class="marketing-shell" style="margin-top:18px;">
            <div class="marketing-card">
                <div class="marketing-kicker">Produto</div>
                <div class="marketing-title" style="font-size:1.75rem;">Apresente valor sem perder precisão.</div>
                <div class="marketing-copy">
                    A primeira impressão é elegante e comercial. A segunda tela é objetiva, operacional e pronta para auditar.
                    Isso permite vender a solução e, no mesmo ambiente, executar a conferência real com segurança.
                </div>
            </div>
            <div class="marketing-card">
                <div class="marketing-kicker">Uso prático</div>
                <div class="marketing-title" style="font-size:1.75rem;">Uma experiência mais moderna e intuitiva.</div>
                <div class="marketing-copy">
                    Cartões com profundidade, tipografia mais forte, áreas de upload mais vivas e uma linguagem visual
                    mais atual, mantendo os resultados já validados exatamente como estão.
                </div>
            </div>
        </div>
        <div class="marketing-grid" style="margin-top:16px;">
            <div class="marketing-feature">
                <small>Conferência</small>
                <b>Diferença exata por campo</b>
                <span>Empresa, motorista, maior diferença, margem GW e observação aparecem no mesmo fluxo de leitura.</span>
            </div>
            <div class="marketing-feature">
                <small>Confiabilidade</small>
                <b>Mesmos números validados</b>
                <span>O visual evolui sem alterar motor, parser, cálculo, tolerância, cruzamento ou regra de status.</span>
            </div>
            <div class="marketing-feature">
                <small>Apresentação</small>
                <b>Marca mais viva</b>
                <span>Sombras, gradientes, volumes e tipografia ajudam o produto a parecer mais premium na demonstração.</span>
            </div>
            <div class="marketing-feature">
                <small>Operação</small>
                <b>Auditoria continua limpa</b>
                <span>A área operacional segue direta: upload, processamento, resumo, gráfico e tabela completa.</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_support_page():
    with st.container(border=True):
        st.markdown(
            """
            <div class="section-head">
                <h3 class="section-title">Suporte Técnico</h3>
                <div class="section-subtitle">Canal rápido para dúvidas operacionais e acompanhamento da apresentação.</div>
            </div>
            <div class="summary-note">
                Em caso de dúvida na leitura dos relatórios, confira o expander <b>Debug da leitura</b> na própria tela de auditoria.
                Para atendimento, utilize o contato interno da operação.
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_news_page():
    with st.container(border=True):
        st.markdown(
            """
            <div class="section-head">
                <h3 class="section-title">Novidades</h3>
                <div class="section-subtitle">Resumo curto das evoluções mais importantes da interface.</div>
            </div>
            <div class="summary-note">
                Novo layout compacto, uploads persistentes, tabela com destaque por campo divergente,
                coluna <b>Margem GW</b>, conferência detalhada, PDF executivo e PDF completo de conferência.
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_footer():
    st.markdown(
        """
        <div class="compact-footer">
            © 2026 FreteVision Logística<br>
            Desenvolvido por Mateus
        </div>
        """,
        unsafe_allow_html=True,
    )


render_app_header()
page = render_compact_menu()

if page == "Visão que move resultados":
    render_marketing_page()

elif page == "Auditoria":
    tolerance = render_tolerance_section()
    file_a, file_b = render_upload_section()

    if "processing" not in st.session_state:
        st.session_state.processing = False
    if "audit_error" not in st.session_state:
        st.session_state.audit_error = None
    if "audit_error_details" not in st.session_state:
        st.session_state.audit_error_details = []
    if "audit_debug" not in st.session_state:
        st.session_state.audit_debug = None

    button_col_1, button_col_2, button_col_3 = st.columns([1.7, 1.5, 1.7])
    with button_col_2:
        clicked = st.button("Iniciar Auditoria", type="primary", use_container_width=True, disabled=st.session_state.processing)

    status_box = st.empty()
    debug_box = st.empty()
    results_box = st.empty()

    if st.session_state.get("audit_debug"):
        with debug_box.container():
            render_audit_debug()

    if st.session_state.audit_error and not st.session_state.processing:
        status_box.error(st.session_state.audit_error)
        for detail in st.session_state.audit_error_details:
            st.error(f"- {detail}")

    if clicked:
        st.session_state.processing = True
        st.session_state.audit_error = None
        st.session_state.audit_error_details = []
        reset_audit_state()
        st.rerun()

    if st.session_state.processing:
        if not file_a or not file_b:
            set_audit_error("Faça upload dos dois relatórios antes de iniciar a auditoria.")

        caminho_a = None
        caminho_b = None
        timings = {}
        started_at = perf_counter()
        try:
            update_processing_view(status_box, 1, 10, "Lendo Relatório A")
            step_started_at = perf_counter()
            caminho_a = salvar_upload_pdf(file_a, "ATUA")
            timings["Salvar upload ATUA"] = perf_counter() - step_started_at
            update_processing_view(status_box, 2, 25, "Lendo Relatório B")
            step_started_at = perf_counter()
            caminho_b = salvar_upload_pdf(file_b, "GW")
            timings["Salvar upload GW"] = perf_counter() - step_started_at

            set_audit_debug(caminho_a=caminho_a, caminho_b=caminho_b, timings=timings)

            with debug_box.container():
                render_audit_debug()

            update_processing_view(status_box, 3, 55, "Cruzando CTEs")
            tolerance_dec = Decimal(str(tolerance)).quantize(Decimal("0.01"))
            step_started_at = perf_counter()
            resultado = auditar(caminho_a, caminho_b, tolerance_dec)
            timings["Leitura e cruzamento"] = perf_counter() - step_started_at
            update_processing_view(status_box, 4, 78, "Aplicando tolerância")
            step_started_at = perf_counter()
            result_df = linhas_para_dataframe(resultado["linhas"])
            result_df = aplicar_margens_gw_visual(result_df, caminho_b)
            result_df = ensure_gw_margin_visual(result_df, caminho_b)
            summary = normalizar_resumo_motor(resultado["resumo"])
            timings["Pós-processamento"] = perf_counter() - step_started_at

            set_audit_debug(
                registros_a=resultado["registros_a"],
                registros_b=resultado["registros_b"],
                caminho_a=caminho_a,
                caminho_b=caminho_b,
                timings=timings,
            )

            with debug_box.container():
                render_audit_debug()

            erros_zero_cte = []
            if not resultado["registros_a"]:
                erros_zero_cte.append("ATUA retornou 0 CTEs.")
            if not resultado["registros_b"]:
                erros_zero_cte.append("GW retornou 0 CTEs.")
            if erros_zero_cte:
                set_audit_error("Falha na leitura dos PDFs.", erros_zero_cte)

            update_processing_view(status_box, 5, 92, "Gerando relatório")
            step_started_at = perf_counter()
            auditoria_io.salvar_historico(get_upload_name(file_a), get_upload_name(file_b), float(tolerance), summary)
            timings["Salvar histórico"] = perf_counter() - step_started_at
            timings["Total"] = perf_counter() - started_at
            update_processing_view(status_box, 5, 100, "Gerando relatório")

            set_audit_debug(
                registros_a=resultado["registros_a"],
                registros_b=resultado["registros_b"],
                caminho_a=caminho_a,
                caminho_b=caminho_b,
                timings=timings,
            )

            st.session_state.update(
                {
                    "df_res": result_df,
                    "resumo": summary,
                    "nome_a": get_upload_name(file_a),
                    "nome_b": get_upload_name(file_b),
                    "tol": tolerance,
                    "caminho_a_temp": caminho_a,
                    "caminho_b_temp": caminho_b,
                }
            )
            clear_export_caches()
            status_box.empty()
        except Exception as exc:
            mensagem = str(exc)
            warn_a = [mensagem] if "ATUA" in mensagem.upper() else []
            warn_b = [mensagem] if "GW" in mensagem.upper() else []
            if not warn_a and not warn_b:
                warn_a = [mensagem]
                warn_b = [mensagem]
            timings["Total até erro"] = perf_counter() - started_at
            set_audit_debug(warn_a=warn_a, warn_b=warn_b, caminho_a=caminho_a, caminho_b=caminho_b, timings=timings)
            set_audit_error(
                titulo_erro_processamento(mensagem),
                [mensagem],
            )

        st.session_state.processing = False
        st.rerun()

    if st.session_state.get("df_res") is not None and not st.session_state["df_res"].empty:
        with results_box.container():
            result = ensure_gw_margin_visual(st.session_state["df_res"], st.session_state.get("caminho_b_temp"))
            if not result.equals(st.session_state["df_res"]):
                st.session_state["df_res"] = result
                clear_export_caches()
            summary = st.session_state["resumo"]
            render_kpis(summary, result)
            render_chart(result)
            render_table(result)
            render_exports(result, summary, st.session_state["nome_a"], st.session_state["nome_b"], st.session_state["tol"])

elif page == "Suporte Técnico":
    render_support_page()
elif page == "Novidades":
    render_news_page()

render_footer()

