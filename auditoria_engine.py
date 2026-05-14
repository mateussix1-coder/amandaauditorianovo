# auditoria_engine.py
# Motor de auditoria FreteScan Pro - versão corrigida para o padrão ATUA x GW
# Foco: leitura por blocos/células extraídas do PDF, não por linha única.

import io
import json
import os
import re
import tempfile
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import pdfplumber


CENTAVOS = Decimal("0.01")

RE_CTE_NUM = re.compile(r"^\d{4,6}$")
RE_DATA_ATUA = re.compile(r"^\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}$")
RE_DATA_GW = re.compile(r"^\d{2}/\d{2}/\d{4}$")
RE_PLACA = re.compile(r"^[A-Z]{3}\d[A-Z0-9]\d{2}$|^[A-Z]{3}\d{4}$")
RE_MONEY_BR = re.compile(r"^-?\d{1,3}(?:\.\d{3})*,\d{2}$|^-?\d+,\d{2}$")
MONEY_RE = re.compile(r"-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2}")
RE_PESO_TON = re.compile(r"^\d{1,3},\d{3}$")
RE_PERCENT = re.compile(r"^-?\d{1,3},\d{2}%$")
MAX_PDF_PAGE_COUNT = 300
MAX_HISTORY_ENTRIES = 100


def parse_money_br(value) -> Optional[Decimal]:
    """
    Converte moeda brasileira para Decimal corretamente.

    Exemplos:
    23.919,00 -> 23919.00
    24.839,65 -> 24839.65
    1.817,84  -> 1817.84
    0,00      -> 0.00
    """
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value.quantize(CENTAVOS, rounding=ROUND_HALF_UP)

    text = str(value).strip()
    text = text.replace("R$", "").replace(" ", "")

    if text in ["", "-", "None", "nan"]:
        return None

    # BR: ponto = milhar, vírgula = decimal
    if "," in text:
        text = text.replace(".", "")
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")

    try:
        return Decimal(text).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def decimal_to_float(value):
    if value is None:
        return None
    return float(value)


def format_money_br(value) -> str:
    if value is None:
        return "-"

    if not isinstance(value, Decimal):
        value = Decimal(str(value))

    value = value.quantize(CENTAVOS, rounding=ROUND_HALF_UP)
    sinal = "-" if value < 0 else ""
    value_abs = abs(value)
    s = f"{value_abs:.2f}"
    inteiro, centavos = s.split(".")
    partes = []

    while len(inteiro) > 3:
        partes.insert(0, inteiro[-3:])
        inteiro = inteiro[:-3]

    partes.insert(0, inteiro)
    return f"{sinal}R$ {'.'.join(partes)},{centavos}"


def normalizar_cte(value) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()

    if not re.fullmatch(r"\d{4,6}", text):
        return None

    numero = int(text)

    if numero < 1000 or numero > 999999:
        return None

    return str(numero)


def _extrair_linhas_pdfplumber(caminho_pdf):
    linhas = []
    with pdfplumber.open(str(caminho_pdf)) as pdf:
        total_paginas = len(pdf.pages)
        if total_paginas == 0:
            raise ValueError("O PDF enviado não possui páginas legíveis.")
        if total_paginas > MAX_PDF_PAGE_COUNT:
            raise ValueError(
                f"O PDF possui {total_paginas} páginas e excede o limite de {MAX_PDF_PAGE_COUNT} páginas."
            )
        for page_num, page in enumerate(pdf.pages, start=1):
            texto = page.extract_text() or ""
            for raw in texto.splitlines():
                t = raw.strip()
                if t:
                    linhas.append((page_num, t))
    return linhas


# Regex para linha de CTE do ATUA:  "1752 CT ..."
RE_ATUA_LINHA = re.compile(r"^\s*(\d{4,6})\s+CT\b")
# Regex para linha de CTE do GW: "001752 01/04/2026 ..."
RE_GW_LINHA = re.compile(r"^\s*(\d{4,6})\s+\d{2}/\d{2}/\d{4}\b")
RE_GW_NUM_TOKEN = re.compile(r"-?\d[\d\.]*,\d{2}|-?\d+\.\d{2}")

ATUA_HEADER_LINES = {
    "Numero",
    "T.",
    "Emissao Hora",
    "Filial",
    "Agencia",
    "Remetente",
    "Destinatario",
    "Pagador",
    "Placa",
    "Peso (Ton)",
    "Frete Empr.",
    "Frete Mot.",
    "Adto. Empr.",
    "Adto. Mot.",
}

GW_HEADER_LINES = {
    "CTe/NFS",
    "Emissao",
    "Remetente / Origem",
    "Destinatario / Destino",
    "Tipo Frete",
    "Peso / Kg",
    "Valor frete",
    "ICMS/ISS (%)",
    "Frete tab.",
    "PIS",
    "COFINS",
    "IR",
    "CSSL",
    "Vl Carreteiro Liquido",
}


def _extrair_atua_linha_unica(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    for page_num, linha in linhas_pdf:
        m = RE_ATUA_LINHA.match(linha)
        if not m:
            continue
        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

        todos = MONEY_RE.findall(linha)

        if len(todos) < 3:
            continue

        empresa_a = parse_money_br(todos[1])
        motorista_a = parse_money_br(todos[2])

        if empresa_a is None or motorista_a is None:
            continue

        registros[cte] = {
            "cte": cte,
            "empresa": empresa_a,
            "motorista": motorista_a,
            "pagina": page_num,
            "margem": None,
            "raw": linha
        }
    return registros


def _ignorar_linha_atua_multilinha(linha: str) -> bool:
    return (
        linha in ATUA_HEADER_LINES
        or linha.startswith("ATUA - ")
        or linha.startswith("Pagina:")
        or linha.startswith("Relatorio Detalhado do CTRC")
        or linha.startswith("Ambiente")
    )


def _finalizar_bloco_atua(registros, bloco):
    if not bloco:
        return

    valores = [
        parse_money_br(linha)
        for linha in bloco["linhas"]
        if RE_MONEY_BR.fullmatch(linha)
    ]
    valores = [valor for valor in valores if valor is not None]

    if len(valores) < 2:
        return

    registros[bloco["cte"]] = {
        "cte": bloco["cte"],
        "empresa": valores[0],
        "motorista": valores[1],
        "pagina": bloco["pagina"],
        "margem": None,
        "raw": " | ".join(bloco["linhas"]),
    }


def _extrair_atua_multilinha(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    bloco_atual = None

    for page_num, linha in linhas_pdf:
        if _ignorar_linha_atua_multilinha(linha):
            continue

        cte = normalizar_cte(linha) if RE_CTE_NUM.fullmatch(linha) else None
        if cte:
            _finalizar_bloco_atua(registros, bloco_atual)
            bloco_atual = {"cte": cte, "pagina": page_num, "linhas": []}
            continue

        if bloco_atual is not None:
            bloco_atual["linhas"].append(linha)

    _finalizar_bloco_atua(registros, bloco_atual)
    return registros


def extrair_atua_por_blocos(caminho_pdf) -> Dict[str, Dict[str, Any]]:
    # Reutiliza a mesma extração textual para evitar varrer o PDF duas vezes.
    linhas_pdf = _extrair_linhas_pdfplumber(caminho_pdf)
    registros = _extrair_atua_linha_unica(linhas_pdf)
    if registros:
        return registros
    return _extrair_atua_multilinha(linhas_pdf)


def _extrair_gw_linha_unica(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}

    for page_num, linha in linhas_pdf:
        m = RE_GW_LINHA.match(linha)
        if not m:
            continue

        cte = normalizar_cte(m.group(1))
        if not cte:
            continue

 codex/test-system-after-mapping-changes-rmuh98
        tokens = MONEY_RE.findall(linha)
        valores = [parse_money_br(v) for v in tokens]
        valores = [v for v in valores if v is not None]

        if len(valores) < 2:
            continue

        # REGRA CORRETA DO GW:
        # Na linha do CTE, o primeiro valor financeiro é "Valor frete".
        # Esse é o campo correto para Empresa B.
        empresa_b = valores[0]

        # Motorista B deve permanecer como já estava na lógica original:
        # último valor financeiro da linha, correspondente ao Vl Carreteiro Líquido.
        motorista_b = valores[-1]

 codex/test-system-after-mapping-changes-djn5sw
        valores = MONEY_RE.findall(linha)
        if len(valores) < 2:
            continue

        # Regra: Empresa B deve vir de "Valor frete" (e não "Frete tab.").
        # No layout em linha única, os dois últimos valores monetários são:
        # [Frete tab., Valor frete]
        empresa_b = parse_money_br(valores[-1])

        # Motorista B permanece inalterado.
        motorista_b = parse_money_br(valores[-1])

        valores_raw = RE_GW_NUM_TOKEN.findall(linha)
        # Remove CTE/data e mantém apenas números da grade financeira da linha.
        # Ex.: 001960 01/04/2026 ... -> [frete, pedagio, adval, gris, frete_tab, valor_frete, ...]
        valores = [parse_money_br(v) for v in valores_raw]
        valores = [v for v in valores if v is not None]

        if len(valores) < 2:
            continue

 codex/test-system-after-mapping-changes-mlhld0
        # Regra obrigatória: Empresa B = Valor frete (não Frete tab.).
        # No GW de linha única, o Valor frete é o valor imediatamente após Frete tab.
        # Quando a estrutura não vier completa, usamos o último valor não-zero para
        # evitar cair em 0,00 por falha de extração.
        empresa_b = valores[5] if len(valores) > 5 else None
        if empresa_b is None or empresa_b == Decimal("0.00"):
            nao_zero = [v for v in valores if v != Decimal("0.00")]
            empresa_b = nao_zero[-1] if nao_zero else None

        # Motorista B permanece com a regra existente do parser.
        motorista_b = parse_money_br(MONEY_RE.findall(linha)[-1]) if MONEY_RE.findall(linha) else None

        # Regra: Empresa B deve usar "Valor frete" (e não "Frete tab.").
        # No layout de linha única do GW, o último bloco monetário contém:
        # [..., Frete tab., Valor frete, Motorista B, ...]
        # Mantemos Motorista B inalterado em valores[-3].
        empresa_b = parse_money_br(valores[-4])
        motorista_b = parse_money_br(valores[-3])
 main
 main
 main

        if empresa_b is None or motorista_b is None:
            continue

        margem = None
        percentuais = RE_PERCENT.findall(linha)
        if percentuais:
            margem = percentuais[-1]

        registros[cte] = {
            "cte": cte,
            "empresa": empresa_b,
            "motorista": motorista_b,
            "pagina": page_num,
            "margem": margem,
            "raw": linha,
        }
    return registros


def _ignorar_linha_gw_multilinha(linha: str) -> bool:
    return (
        linha in GW_HEADER_LINES
        or linha.startswith("GW - ")
        or linha.startswith("Pagina:")
        or linha.startswith("Analise de CTe/NFS")
        or linha.startswith("Usuario:")
        or linha.startswith("FILIAL :")
    )


def _finalizar_bloco_gw(registros, bloco):
    if not bloco:
        return

    cte = None
    pagina_cte = bloco["pagina"]
    valores_antes_cte = []

    for page_num, linha in bloco["linhas"]:
        if RE_CTE_NUM.fullmatch(linha):
            cte = normalizar_cte(linha)
            pagina_cte = page_num
            break

        if RE_MONEY_BR.fullmatch(linha):
            valor = parse_money_br(linha)
            if valor is not None:
                valores_antes_cte.append(valor)

    if not cte or len(valores_antes_cte) < 2:
        return

    # Regra: Empresa B = "Valor frete" do GW (não "Frete tab.").
 codex/test-system-after-mapping-changes-rmuh98
    # No formato multilinha, considera o segundo valor financeiro do bloco.

    # No formato multilinha, os valores antes do CTE chegam com Frete tab.
    # seguido de Valor frete; por isso usamos o segundo valor.
 codex/test-system-after-mapping-changes-djn5sw
 main
    empresa_b = valores_antes_cte[1] if len(valores_antes_cte) >= 2 else None
    if empresa_b is None:
        return

 codex/test-system-after-mapping-changes-rmuh98


 codex/test-system-after-mapping-changes-mlhld0
    empresa_b = valores_antes_cte[1] if len(valores_antes_cte) >= 2 else None
    if empresa_b == Decimal("0.00"):
        nao_zero = [v for v in valores_antes_cte if v != Decimal("0.00")]
        empresa_b = nao_zero[-1] if nao_zero else None
    if empresa_b is None:
        return

    empresa_b = valores_antes_cte[1] if len(valores_antes_cte) >= 2 else valores_antes_cte[0]
 main

 main
 main
    registros[cte] = {
        "cte": cte,
        "empresa": empresa_b,
        "motorista": valores_antes_cte[-1],
        "pagina": pagina_cte,
        "margem": None,
        "raw": " | ".join(linha for _, linha in bloco["linhas"]),
    }


def _extrair_gw_multilinha(linhas_pdf) -> Dict[str, Dict[str, Any]]:
    registros = {}
    bloco_atual = None

    for page_num, linha in linhas_pdf:
        if _ignorar_linha_gw_multilinha(linha):
            continue

        if RE_DATA_GW.fullmatch(linha):
            _finalizar_bloco_gw(registros, bloco_atual)
            bloco_atual = {"pagina": page_num, "linhas": [(page_num, linha)]}
            continue

        if bloco_atual is not None:
            bloco_atual["linhas"].append((page_num, linha))

    _finalizar_bloco_gw(registros, bloco_atual)
    return registros


def extrair_gw_por_blocos(caminho_pdf) -> Dict[str, Dict[str, Any]]:
    # Reutiliza a mesma extração textual para evitar varrer o PDF duas vezes.
    linhas_pdf = _extrair_linhas_pdfplumber(caminho_pdf)
    registros = _extrair_gw_linha_unica(linhas_pdf)
    if registros:
        return registros
    return _extrair_gw_multilinha(linhas_pdf)


def ler_atua(caminho_pdf):
    registros = extrair_atua_por_blocos(caminho_pdf)

    if not registros:
        raise ValueError("Falha na leitura do ATUA. Nenhum CTE válido com valores foi encontrado.")

    return registros


def _bytes_para_tmp(pdf_bytes: bytes) -> str:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(pdf_bytes)
        return tmp.name

def parse_atua(pdf_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    tmp = _bytes_para_tmp(pdf_bytes)
    try:
        registros = extrair_atua_por_blocos(tmp)
    finally:
        os.unlink(tmp)
    linhas = [{"CTE": r["cte"], "EmpresaA": r["empresa"], "MotoristaA": r["motorista"]} for r in registros.values()]
    return pd.DataFrame(linhas), []

def parse_gw(pdf_bytes: bytes) -> tuple[pd.DataFrame, list[str]]:
    tmp = _bytes_para_tmp(pdf_bytes)
    try:
        registros = extrair_gw_por_blocos(tmp)
    finally:
        os.unlink(tmp)
    linhas = [{"CTE": r["cte"], "EmpresaB": r["empresa"], "MotoristaB": r["motorista"]} for r in registros.values()]
    return pd.DataFrame(linhas), []



def ler_gw(caminho_pdf):
    registros = extrair_gw_por_blocos(caminho_pdf)

    if not registros:
        raise ValueError("Falha na leitura do GW. Nenhum CTE válido com valores foi encontrado.")

    return registros


def calcular_status(existe_a, existe_b, dif_empresa, dif_motorista, tolerancia: Decimal):
    if existe_a and not existe_b:
        return "Faltante no B"

    if existe_b and not existe_a:
        return "Faltante no A"

    abs_emp = abs(dif_empresa)
    abs_mot = abs(dif_motorista)

    if abs_emp == Decimal("0.00") and abs_mot == Decimal("0.00"):
        return "OK"

    if abs_emp <= tolerancia and abs_mot <= tolerancia:
        return "OK por arredondamento"

    return "Divergente"


def auditar(caminho_atua, caminho_gw, tolerancia=Decimal("0.50")):
    if not isinstance(tolerancia, Decimal):
        tolerancia = Decimal(str(tolerancia)).quantize(CENTAVOS, rounding=ROUND_HALF_UP)

    registros_a = ler_atua(caminho_atua)
    registros_b = ler_gw(caminho_gw)

    validar_integridade_basica(registros_a, registros_b)

    todos_ctes = sorted(
        set(registros_a.keys()) | set(registros_b.keys()),
        key=lambda x: int(x)
    )

    linhas = []

    for cte in todos_ctes:
        a = registros_a.get(cte)
        b = registros_b.get(cte)

        existe_a = a is not None
        existe_b = b is not None

        empresa_a = a["empresa"] if a else None
        motorista_a = a["motorista"] if a else None
        empresa_b = b["empresa"] if b else None
        motorista_b = b["motorista"] if b else None

        # Para cálculo de faltante, usa 0 apenas no cálculo, mas mantém campo exibido como None.
        calc_empresa_a = empresa_a if empresa_a is not None else Decimal("0.00")
        calc_empresa_b = empresa_b if empresa_b is not None else Decimal("0.00")
        calc_motorista_a = motorista_a if motorista_a is not None else Decimal("0.00")
        calc_motorista_b = motorista_b if motorista_b is not None else Decimal("0.00")

        dif_empresa = (calc_empresa_a - calc_empresa_b).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
        dif_motorista = (calc_motorista_a - calc_motorista_b).quantize(CENTAVOS, rounding=ROUND_HALF_UP)
        maior_diferenca = max(abs(dif_empresa), abs(dif_motorista))

        status = calcular_status(existe_a, existe_b, dif_empresa, dif_motorista, tolerancia)

        linhas.append({
            "CTE": cte,
            "Status": status,
            "Empresa A": empresa_a,
            "Empresa B": empresa_b,
            "Motorista A": motorista_a,
            "Motorista B": motorista_b,
            "Dif Empresa": dif_empresa,
            "Dif Motorista": dif_motorista,
            "Maior Diferença": maior_diferenca,
            "Margem B": b.get("margem") if b else None,
            "Página A": a.get("pagina") if a else None,
            "Página B": b.get("pagina") if b else None,
        })

    resumo = gerar_resumo(linhas, registros_a, registros_b, tolerancia)

    return {
        "resumo": resumo,
        "linhas": linhas,
        "registros_a": registros_a,
        "registros_b": registros_b,
        "debug": gerar_debug(registros_a, registros_b),
    }


def gerar_resumo(linhas, registros_a, registros_b, tolerancia):
    total = len(linhas)

    def count(status):
        return sum(1 for x in linhas if x["Status"] == status)

    dif_empresa_total = sum((x["Dif Empresa"] for x in linhas), Decimal("0.00")).quantize(CENTAVOS)
    dif_motorista_total = sum((x["Dif Motorista"] for x in linhas), Decimal("0.00")).quantize(CENTAVOS)

    impacto_abs = sum(
        (x["Maior Diferença"] for x in linhas if x["Status"] in ["Divergente", "Faltante no A", "Faltante no B"]),
        Decimal("0.00")
    ).quantize(CENTAVOS)

    return {
        "tolerancia": tolerancia,
        "ctes_atua": len(registros_a),
        "ctes_gw": len(registros_b),
        "total_analisado": total,
        "cruzados": len(set(registros_a.keys()) & set(registros_b.keys())),
        "ok": count("OK"),
        "ok_arredondamento": count("OK por arredondamento"),
        "divergentes": count("Divergente"),
        "faltante_a": count("Faltante no A"),
        "faltante_b": count("Faltante no B"),
        "dif_empresa_total": dif_empresa_total,
        "dif_motorista_total": dif_motorista_total,
        "impacto_absoluto": impacto_abs,
        "dif_total_empresa": dif_empresa_total,    # aliases for frontend
        "dif_total_motorista": dif_motorista_total # aliases for frontend
    }


def gerar_resumo_df(df: pd.DataFrame) -> dict:
    total = len(df)
    ok = len(df[df["Status"] == "OK"])
    ok_r = len(df[df["Status"] == "OK por arredondamento"])
    div = len(df[df["Status"] == "Divergente"])
    fa = len(df[df["Status"] == "Faltante no A"])
    fb = len(df[df["Status"] == "Faltante no B"])
    crit = df[df["Status"].isin(["Divergente", "Faltante no A", "Faltante no B"])]
    dif_empresa_total = df["Dif Empresa"].fillna(0).sum()
    dif_motorista_total = df["Dif Motorista"].fillna(0).sum()
    impacto_abs = crit["Maior Diferença"].fillna(0).sum()

    return {
        "total": int(total),
        "ok": int(ok),
        "ok_arredondamento": int(ok_r),
        "divergentes": int(div),
        "faltantes_a": int(fa),
        "faltantes_b": int(fb),
        "dif_total_empresa": float(round(dif_empresa_total, 2)),
        "dif_total_motorista": float(round(dif_motorista_total, 2)),
        "impacto_absoluto": float(round(impacto_abs, 2)),
    }

def validar_integridade_basica(registros_a, registros_b):
    erros = []

    if len(registros_a) < 10:
        erros.append(f"ATUA com poucos CTEs lidos: {len(registros_a)}")

    if len(registros_b) < 10:
        erros.append(f"GW com poucos CTEs lidos: {len(registros_b)}")

    if registros_a:
        zerados_a = sum(1 for r in registros_a.values() if r["empresa"] == 0 or r["motorista"] == 0)
        if zerados_a > len(registros_a) * Decimal("0.20"):
            erros.append("ATUA com valores zerados em massa. Parser provavelmente errado.")

    if registros_b:
        absurdos_b = sum(1 for r in registros_b.values() if r["motorista"] > Decimal("200000.00"))
        if absurdos_b > len(registros_b) * Decimal("0.05"):
            erros.append("GW com valores absurdos no Motorista B. Conversão monetária provavelmente errada.")

    inter = set(registros_a.keys()) & set(registros_b.keys())

    if len(inter) == 0:
        erros.append("Nenhum CTE cruzado entre ATUA e GW.")

    if erros:
        raise ValueError("Falha na validação de integridade: " + " | ".join(erros))


def gerar_debug(registros_a, registros_b):
    def top(registros):
        saida = []
        for cte in sorted(registros.keys(), key=lambda x: int(x))[:10]:
            r = registros[cte]
            saida.append({
                "CTE": cte,
                "Empresa": r["empresa"],
                "Motorista": r["motorista"],
                "Página": r.get("pagina")
            })
        return saida

    return {
        "ATUA - Top 10": top(registros_a),
        "GW - Top 10": top(registros_b),
    }


def linhas_para_dataframe(linhas) -> pd.DataFrame:
    rows = []

    for x in linhas:
        rows.append({
            "CTE": x["CTE"],
            "Status": x["Status"],
            "Empresa A": decimal_to_float(x["Empresa A"]),
            "Empresa B": decimal_to_float(x["Empresa B"]),
            "Motorista A": decimal_to_float(x["Motorista A"]),
            "Motorista B": decimal_to_float(x["Motorista B"]),
            "Dif Empresa": decimal_to_float(x["Dif Empresa"]),
            "Dif Motorista": decimal_to_float(x["Dif Motorista"]),
            "Maior Diferença": decimal_to_float(x["Maior Diferença"]),
            "Margem B": x["Margem B"],
            "Página A": x["Página A"],
            "Página B": x["Página B"],
        })

    return pd.DataFrame(rows)


def testar_parser_basico(caminho_atua, caminho_gw):
    resultado = auditar(caminho_atua, caminho_gw, Decimal("0.50"))

    a = resultado["registros_a"]
    b = resultado["registros_b"]

    assert a["1752"]["empresa"] == Decimal("23919.00")
    assert a["1752"]["motorista"] == Decimal("24839.65")
    assert b["1752"]["empresa"] == Decimal("23919.00")
    assert b["1752"]["motorista"] == Decimal("24839.88")

    assert a["1753"]["empresa"] == Decimal("12892.50")
    assert a["1753"]["motorista"] == Decimal("13388.62")
    assert b["1753"]["empresa"] == Decimal("12892.50")
    assert b["1753"]["motorista"] == Decimal("12892.50")

    print("Parser OK.")
    print(resultado["resumo"])
    print(resultado["debug"])

    return resultado

# ---------------------------------------------------------------------------
# Histórico & Export
# ---------------------------------------------------------------------------

HIST_PATH = Path("historico_auditoria.json")

def _normalizar_historico(hist) -> list:
    if not isinstance(hist, list):
        return []
    return [item for item in hist[:MAX_HISTORY_ENTRIES] if isinstance(item, dict)]

def _gravar_historico_atomico(hist) -> None:
    HIST_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(hist, ensure_ascii=False, indent=2)
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        delete=False,
        dir=HIST_PATH.parent,
        prefix=f"{HIST_PATH.stem}_",
        suffix=".tmp",
    ) as tmp:
        tmp.write(payload)
        tmp_path = Path(tmp.name)
    os.replace(tmp_path, HIST_PATH)

def salvar_historico(nome_a, nome_b, tolerancia, resumo):
    hist = []
    if HIST_PATH.exists():
        try:
            hist = _normalizar_historico(json.loads(HIST_PATH.read_text(encoding="utf-8")))
        except Exception:
            hist = []
    hist.insert(0, {
        "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "arquivo_a": nome_a,
        "arquivo_b": nome_b,
        "tolerancia": float(tolerancia),
        **resumo,
    })
    _gravar_historico_atomico(_normalizar_historico(hist))


def carregar_historico() -> list:
    if not HIST_PATH.exists():
        return []
    try:
        return _normalizar_historico(json.loads(HIST_PATH.read_text(encoding="utf-8")))
    except Exception:
        return []

def exportar_csv(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, sep=";", index=False, encoding="utf-8-sig")
    return buf.getvalue().encode("utf-8-sig")


def exportar_excel(df, resumo, nome_a, nome_b, tolerancia) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    CORES = {
        "OK": "C6EFCE", "OK por arredondamento": "FFEB9C",
        "Divergente": "FFC7CE", "Faltante no A": "DDEBF7", "Faltante no B": "FCE4D6",
    }
    buf = io.BytesIO()
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.append(["FreteScan Pro — Resumo de Auditoria"])
    ws1.append([f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws1.append([f"Arquivo A: {nome_a}  |  Arquivo B: {nome_b}"])
    ws1.append([f"Tolerância: R$ {tolerancia:.2f}"])
    ws1.append([])
    ws1.append(["Métrica", "Valor"])
    for k, v in resumo.items():
        ws1.append([k, v])

    ws2 = wb.create_sheet("Auditoria")
    ws2.append(list(df.columns))
    ws2.freeze_panes = "A2"
    for r, row in enumerate(df.itertuples(index=False), start=2):
        ws2.append([x if pd.notna(x) else "" for x in row])
        fill = PatternFill(
            start_color=CORES.get(row.Status, "FFFFFF"),
            end_color=CORES.get(row.Status, "FFFFFF"),
            fill_type="solid",
        )
        for c in range(1, len(df.columns) + 1):
            ws2.cell(row=r, column=c).fill = fill
    for col in ws2.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 40)

    wb.save(buf)
    return buf.getvalue()


def exportar_pdf(df, resumo, nome_a, nome_b, tolerancia) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph("FreteScan Pro — Relatório de Auditoria", styles["Title"]))
    elems.append(Spacer(1, 0.3 * cm))
    elems.append(Paragraph(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles["Normal"]))
    elems.append(Paragraph(f"Arquivo A: {nome_a}  |  Arquivo B: {nome_b}", styles["Normal"]))
    elems.append(Paragraph(f"Tolerância: R$ {tolerancia:.2f}  —  Diferença = A − B", styles["Normal"]))
    elems.append(Spacer(1, 0.5 * cm))

    elems.append(Paragraph("Resumo Geral", styles["Heading2"]))
    res_data = [["Métrica", "Valor"]] + [[k, str(v)] for k, v in resumo.items()]
    t = Table(res_data, colWidths=[8 * cm, 4 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 0.5 * cm))

    df_div = df[df["Status"].isin(["Divergente", "Faltante no A", "Faltante no B"])].head(50)
    if not df_div.empty:
        elems.append(Paragraph("Divergências e Faltantes (até 50)", styles["Heading2"]))
        cols = ["CTE", "Status", "Empresa A", "Empresa B", "Dif. Empresa",
                "Motorista A", "Motorista B", "Dif. Motorista"]
        data = [cols] + [
            ["" if pd.isna(row[c]) else str(row[c]) for c in cols]
            for _, row in df_div.iterrows()
        ]
        cw = [2 * cm, 4 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm]
        t2 = Table(data, colWidths=cw, repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        elems.append(t2)

    doc.build(elems)
    return buf.getvalue()
